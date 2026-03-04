# SPDX-FileCopyrightText: 2025 QinHan
# SPDX-License-Identifier: MPL-2.0
import asyncio
import os
import tempfile
from dataclasses import dataclass
from io import BytesIO
from typing import Self, Literal, List, Optional

import openpyxl
from openpyxl.cell import Cell

from collabtrans.agents.segments_agent import SegmentsTranslateAgentConfig, SegmentsTranslateAgent
from collabtrans.ir.document import Document
from collabtrans.translator.ai_translator.base import AiTranslatorConfig, AiTranslator
from collabtrans.utils.memory_utils import log_memory


@dataclass
class XlsxTranslatorConfig(AiTranslatorConfig):
    insert_mode: Literal["replace", "append", "prepend"] = "replace"
    separator: str = "\n"
    # Specify translation region list.
    # Example: ["Sheet1!A1:B10", "C:D", "E5"]
    # If sheet name is not specified (like "C:D"), it applies to all sheets.
    # If None or empty list, translate all text in the entire file.
    translate_regions: Optional[List[str]] = None


class XlsxTranslator(AiTranslator):
    def __init__(self, config: XlsxTranslatorConfig):
        super().__init__(config=config)
        self.chunk_size = config.chunk_size
        self.translate_agent = None
        if not self.skip_translate:
            agent_config = SegmentsTranslateAgentConfig(
                custom_prompt=config.custom_prompt,
                to_lang=config.to_lang,
                base_url=config.base_url,
                api_key=config.api_key,
                api_type=getattr(config, 'api_type', 'openai'),
                model_id=config.model_id,
                temperature=config.temperature,
                thinking=config.thinking,
                concurrent=config.concurrent,
                timeout=config.timeout,
                logger=self.logger,
                glossary_dict=config.glossary_dict,
                retry=config.retry
            )
            self.translate_agent = SegmentsTranslateAgent(agent_config)
        self.insert_mode = config.insert_mode
        self.separator = config.separator
        # --- New features ---
        self.translate_regions = config.translate_regions

    def _pre_translate(self, document: Document):
        """
        Preprocess XLSX file in low-memory mode:
        - Use read_only workbook to scan and collect cells to translate.
        - Do NOT keep workbook object in memory; only return coordinates and texts.

        Returns:
            cells_to_translate: list of {"sheet_name": str, "coordinate": str}
            original_texts: list of cell text (same order as cells_to_translate)
        """
        # Use read_only + data_only to reduce memory while scanning
        workbook = openpyxl.load_workbook(BytesIO(document.content), read_only=True, data_only=True)
        log_memory(self.logger, "xlsx: after load_workbook (read_only)", f"file size {len(document.content) / (1024*1024):.2f} MB")
        cells_to_translate: list[dict] = []
        original_texts: list[str] = []

        # --- Step 1: Collect text cells that need translation based on whether regions are specified ---
        if not self.translate_regions:  # Also handle None or empty list cases
            # No regions: translate all string cells
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.data_type == "s":
                            cells_to_translate.append({
                                "sheet_name": sheet.title,
                                "coordinate": cell.coordinate,
                            })
                            original_texts.append(cell.value)
        else:
            # If translation regions are specified, only search within these regions
            processed_coordinates = set()
            regions_by_sheet: dict[str, list[str]] = {}
            all_sheet_regions: list[str] = []
            for region in self.translate_regions:
                if "!" in region:
                    sheet_name, cell_range = region.split("!", 1)
                    if sheet_name not in regions_by_sheet:
                        regions_by_sheet[sheet_name] = []
                    regions_by_sheet[sheet_name].append(cell_range)
                else:
                    all_sheet_regions.append(region)

            for sheet in workbook.worksheets:
                sheet_specific_ranges = regions_by_sheet.get(sheet.title, [])
                total_ranges_for_this_sheet = sheet_specific_ranges + all_sheet_regions

                if not total_ranges_for_this_sheet:
                    continue

                for cell_range in total_ranges_for_this_sheet:
                    try:
                        cells_in_range = sheet[cell_range]

                        # Flatten to 1D list regardless of whether it returns single cell, 1D tuple or 2D tuple
                        flat_cells: list[Cell] = []
                        if isinstance(cells_in_range, Cell):
                            flat_cells.append(cells_in_range)
                        elif isinstance(cells_in_range, tuple):
                            for item in cells_in_range:
                                if isinstance(item, Cell):
                                    flat_cells.append(item)
                                elif isinstance(item, tuple):
                                    for cell in item:
                                        flat_cells.append(cell)

                        for cell in flat_cells:
                            full_coordinate = (sheet.title, cell.coordinate)
                            if full_coordinate in processed_coordinates:
                                continue

                            if isinstance(cell.value, str) and cell.data_type == "s":
                                cells_to_translate.append({
                                    "sheet_name": sheet.title,
                                    "coordinate": cell.coordinate,
                                })
                                original_texts.append(cell.value)
                                processed_coordinates.add(full_coordinate)
                    except Exception as e:
                        self.logger.warning(f"Skipping invalid range '{cell_range}' in worksheet '{sheet.title}'. Error: {e}")

        workbook.close()
        log_memory(self.logger, "xlsx: after collecting cells (read_only)", f"{len(cells_to_translate)} cells")
        return cells_to_translate, original_texts

    def _after_translate(self, document: Document, cells_to_translate, translated_texts, original_texts):
        # Load a writable workbook from original document bytes for write-back
        workbook = openpyxl.load_workbook(BytesIO(document.content))
        log_memory(self.logger, "xlsx: _after_translate load_workbook (writable)", f"{len(cells_to_translate)} cells")
        for i, cell_info in enumerate(cells_to_translate):
            sheet_name = cell_info["sheet_name"]
            coordinate = cell_info["coordinate"]
            translated_text = translated_texts[i]
            original_text = original_texts[i]

            # Locate worksheet and cell
            sheet = workbook[sheet_name]
            if self.insert_mode == "replace":
                sheet[coordinate] = translated_text
            elif self.insert_mode == "append":
                sheet[coordinate] = original_text + self.separator + translated_text
            elif self.insert_mode == "prepend":
                sheet[coordinate] = translated_text + self.separator + original_text
            else:
                self.logger.error("Invalid XlsxTranslatorConfig parameter")

        # Save to a temp file instead of BytesIO to reduce peak memory for large workbooks.
        # workbook.save(BytesIO()) + getvalue() keeps both the full buffer and a bytes copy in memory,
        # which can trigger OOM when the xlsx is large (e.g. 10MB+ with many cells).
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        try:
            os.close(fd)
            workbook.save(path)
            with open(path, "rb") as f:
                return f.read()
        finally:
            workbook.close()
            try:
                os.unlink(path)
            except OSError:
                pass

    def translate(self, document: Document) -> Self:

        cells_to_translate, original_texts = self._pre_translate(document)
        if not cells_to_translate:
            print("\nNo plain text content found in specified regions that needs translation.")
            return self
        if self.glossary_agent:
            self.glossary_dict_gen = self.glossary_agent.send_segments(original_texts, self.chunk_size)
            if self.translate_agent:
                self.translate_agent.update_glossary_dict(self.glossary_dict_gen)
        # --- Step 2: Call translation function ---
        if self.translate_agent:
            translated_texts = self.translate_agent.send_segments(original_texts, self.chunk_size)
        else:
            translated_texts = original_texts

        document.content = self._after_translate(document, cells_to_translate, translated_texts, original_texts)
        return self

    async def translate_async(self, document: Document) -> Self:

        cells_to_translate, original_texts = await asyncio.to_thread(self._pre_translate, document)
        log_memory(self.logger, "xlsx: after _pre_translate (read_only)", f"{len(cells_to_translate)} cells")
        if not cells_to_translate:
            print("\nNo plain text content found in specified regions that needs translation.")
            return self

        if self.glossary_agent:
            self.glossary_dict_gen = await self.glossary_agent.send_segments_async(original_texts, self.chunk_size)
            if self.translate_agent:
                self.translate_agent.update_glossary_dict(self.glossary_dict_gen)

        # --- Step 2: Call translation function ---
        log_memory(self.logger, "xlsx: before send_segments_async", f"{len(original_texts)} segments")
        if self.translate_agent:
            translated_texts = await self.translate_agent.send_segments_async(original_texts, self.chunk_size)
        else:
            translated_texts = original_texts
        log_memory(self.logger, "xlsx: after send_segments_async", f"{len(translated_texts)} results")
        log_memory(self.logger, "xlsx: before _after_translate (write back)", "")
        document.content = await asyncio.to_thread(self._after_translate, document, cells_to_translate,
                                                   translated_texts, original_texts)
        log_memory(self.logger, "xlsx: after _after_translate", f"result size {len(document.content) / (1024*1024):.2f} MB")
        return self
