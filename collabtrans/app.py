# SPDX-FileCopyrightText: 2025 QinHan
# SPDX-License-Identifier: MPL-2.0
import asyncio
import base64
import binascii
import json
import logging
import os
import shutil
import socket
import tempfile
import time
import uuid
from contextlib import asynccontextmanager, closing
from pathlib import Path
from typing import List, Dict, Any, Optional, Literal, Union, Annotated, TYPE_CHECKING, Type

import httpx
import uvicorn
from fastapi import FastAPI, HTTPException, APIRouter, Body, Path as FastApiPath, Request, UploadFile, File, Form
from fastapi.openapi.docs import get_swagger_ui_html, get_swagger_ui_oauth2_redirect_html, get_redoc_html
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field, field_validator, model_validator, AliasChoices, ValidationError, TypeAdapter

from collabtrans import __version__
# Initialize project logging (save to logs/app.log and output to console)
import collabtrans.logger.logger  # noqa: F401

# Module logger
logger = logging.getLogger(__name__)
from collabtrans.agents.agent import ThinkingMode
from collabtrans.agents.glossary_agent import GlossaryAgentConfig
from collabtrans.exporter.md.types import ConvertEngineType
# --- Core code imports ---
from collabtrans.global_values.conditional_import import DOCLING_EXIST
from collabtrans.workflow.base import Workflow
from collabtrans.workflow.docx_workflow import DocxWorkflow, DocxWorkflowConfig
from collabtrans.workflow.epub_workflow import EpubWorkflow, EpubWorkflowConfig
# --- HTML WORKFLOW IMPORT START ---
from collabtrans.workflow.html_workflow import HtmlWorkflow, HtmlWorkflowConfig
# --- HTML WORKFLOW IMPORT END ---
from collabtrans.workflow.interfaces import DocxExportable, EpubExportable
from collabtrans.workflow.interfaces import HTMLExportable, MDFormatsExportable, TXTExportable, JsonExportable, \
    XlsxExportable, SrtExportable, CsvExportable
from collabtrans.workflow.json_workflow import JsonWorkflow, JsonWorkflowConfig
from collabtrans.workflow.md_based_workflow import MarkdownBasedWorkflow, MarkdownBasedWorkflowConfig
from collabtrans.workflow.srt_workflow import SrtWorkflow, SrtWorkflowConfig
from collabtrans.workflow.txt_workflow import TXTWorkflow, TXTWorkflowConfig
from collabtrans.workflow.xlsx_workflow import XlsxWorkflow, XlsxWorkflowConfig

if DOCLING_EXIST or TYPE_CHECKING:
    from collabtrans.converter.x2md.converter_docling import ConverterDoclingConfig
from collabtrans.converter.x2md.converter_mineru import ConverterMineruConfig
from collabtrans.exporter.md.md2html_exporter import MD2HTMLExporterConfig
from collabtrans.exporter.txt.txt2html_exporter import TXT2HTMLExporterConfig

# --- Authentication module imports ---
try:
    from collabtrans.auth import AuthConfig, AuthMiddleware, auth_router, auth_compat_router, init_auth
    AUTH_AVAILABLE = True
    print(f"Authentication module imported successfully")
except ImportError as e:
    AUTH_AVAILABLE = False
    print(f"Warning: Authentication module unavailable, skipping auth features. Error: {e}")
except Exception as e:
    AUTH_AVAILABLE = False
    print(f"Warning: Authentication module initialization failed, skipping auth features. Error: {e}")
# --- Authentication module imports END ---
from collabtrans.translator.ai_translator.md_translator import MDTranslatorConfig
from collabtrans.translator.ai_translator.txt_translator import TXTTranslatorConfig
from collabtrans.translator.ai_translator.json_translator import JsonTranslatorConfig
from collabtrans.exporter.js.json2html_exporter import Json2HTMLExporterConfig
from collabtrans.translator.ai_translator.xlsx_translator import XlsxTranslatorConfig
from collabtrans.exporter.xlsx.xlsx2html_exporter import Xlsx2HTMLExporterConfig
from collabtrans.translator.ai_translator.docx_translator import DocxTranslatorConfig
from collabtrans.exporter.docx.docx2html_exporter import Docx2HTMLExporterConfig
from collabtrans.translator.ai_translator.srt_translator import SrtTranslatorConfig
from collabtrans.exporter.srt.srt2html_exporter import Srt2HTMLExporterConfig
from collabtrans.translator.ai_translator.epub_translator import EpubTranslatorConfig
from collabtrans.exporter.epub.epub2html_exporter import Epub2HTMLExporterConfig
# --- HTML TRANSLATOR IMPORT START ---
from collabtrans.translator.ai_translator.html_translator import HtmlTranslatorConfig
# --- HTML TRANSLATOR IMPORT END ---
# ------------------------------------

from collabtrans.logger import global_logger
from collabtrans.translator import default_params
from collabtrans.utils.resource_utils import resource_path
from collabtrans.config.env_detector import is_production, get_config_path, get_dev_config_path, get_prod_config_path

# --- Optional: Playwright for server-side PDF export ---
try:
    from playwright.async_api import async_playwright
    PLAYWRIGHT_AVAILABLE = True
except Exception:
    PLAYWRIGHT_AVAILABLE = False


pdf_router = APIRouter()
preview_router = APIRouter()


class PdfExportRequest(BaseModel):
    html_url: str
    file_name: str | None = None
    # Page settings (optional)
    format: str | None = "A4"
    margin_top: str | None = "10mm"
    margin_right: str | None = "10mm"
    margin_bottom: str | None = "10mm"
    margin_left: str | None = "10mm"


@pdf_router.post("/export/pdf")
async def export_pdf(req: PdfExportRequest, request: Request):
    if not PLAYWRIGHT_AVAILABLE:
        raise HTTPException(status_code=503, detail="Playwright not installed, cannot generate PDF. Please install optional dependency 'pdf_export'.")

    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            context = await browser.new_context()
            # Inject current request session cookies into headless browser to access protected pages
            try:
                host = request.headers.get('host', '')
                cookie_domain = host.split(':')[0] if host else 'localhost'
                cookies_to_set = []
                for name, value in request.cookies.items():
                    cookies_to_set.append({
                        'name': name,
                        'value': value,
                        'domain': cookie_domain,
                        'path': '/',
                        'httpOnly': False,
                        'secure': False,
                        'sameSite': 'Lax',
                    })
                if cookies_to_set:
                    await context.add_cookies(cookies_to_set)
            except Exception as _:
                pass

            page = await context.new_page()
            # Normalize URL: support frontend passing relative paths starting with "/"
            html_url = req.html_url
            if html_url.startswith('/'):
                origin = f"{request.url.scheme}://{request.headers.get('host')}"
                html_url = origin + html_url
            # If the passed URL is a download trigger interface (like /service/download/.../html), direct navigation will trigger download instead of display.
            # Here we change to fetch HTML text on the backend and render with set_content.
            try:
                # Construct Cookie header, reuse request session
                cookie_header = "; ".join([f"{k}={v}" for k, v in request.cookies.items()]) if request.cookies else ""
                headers = {"Cookie": cookie_header} if cookie_header else {}
                resp = await httpx_client.get(html_url, headers=headers)
                resp.raise_for_status()
                html_text = resp.text
            except Exception as fe:
                raise HTTPException(status_code=502, detail=f"Failed to get HTML: {fe}")

            # Inject <base> to ensure relative resources (CSS/JS/images) are parsed correctly
            from urllib.parse import urlparse
            parsed = urlparse(html_url)
            origin = f"{parsed.scheme}://{parsed.netloc}"
            base_href = origin + "/"
            if "<head" in html_text:
                html_with_base = html_text.replace("<head>", f"<head><base href=\"{base_href}\">", 1)
            else:
                html_with_base = f"<head><base href=\"{base_href}\"></head>" + html_text

            await page.set_content(html_with_base, wait_until="networkidle")
            # Generate PDF (no headers/footers, preserve background, use CSS page size)
            pdf_bytes = await page.pdf(
                format=req.format or "A4",
                print_background=True,
                display_header_footer=False,
                prefer_css_page_size=True,
                margin={
                    "top": req.margin_top or "10mm",
                    "right": req.margin_right or "10mm",
                    "bottom": req.margin_bottom or "10mm",
                    "left": req.margin_left or "10mm",
                }
            )
            await context.close()
            await browser.close()

            filename = (req.file_name or "document") + ".pdf"
            headers = {
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
            }
            return Response(content=pdf_bytes, media_type="application/pdf", headers=headers)
    except Exception as e:
        logger.error(f"Failed to export PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to export PDF: {e}")


class PdfExportHtmlRequest(BaseModel):
    html: str
    file_name: str | None = None
    base_url: str | None = None
    format: str | None = "A4"
    margin_top: str | None = "10mm"
    margin_right: str | None = "10mm"
    margin_bottom: str | None = "10mm"
    margin_left: str | None = "10mm"


@pdf_router.post("/export/pdf/from-html")
async def export_pdf_from_html(req: PdfExportHtmlRequest):
    if not PLAYWRIGHT_AVAILABLE:
        raise HTTPException(status_code=503, detail="Playwright not installed, cannot generate PDF. Please install optional dependency 'pdf_export'.")
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            context = await browser.new_context()
            page = await context.new_page()

            html_text = req.html
            base_href = (req.base_url or '').rstrip('/') + '/'
            
            # Enhance HTML content to ensure images and tables display correctly
            enhanced_html = html_text
            
            # Add base tag to ensure relative paths are parsed correctly
            if base_href and "<head" in enhanced_html:
                enhanced_html = enhanced_html.replace("<head>", f"<head><base href=\"{base_href}\">", 1)
            elif base_href:
                enhanced_html = f"<head><base href=\"{base_href}\"></head>" + enhanced_html
            
            # Add CSS to ensure images and tables display correctly in PDF
            css_enhancement = """
            <style>
                /* Ensure images display correctly in PDF */
                img {
                    max-width: 100% !important;
                    height: auto !important;
                    page-break-inside: avoid;
                }
                
                /* Ensure tables display correctly in PDF */
                table {
                    width: 100% !important;
                    border-collapse: collapse !important;
                    page-break-inside: avoid;
                }
                
                table, th, td {
                    border: 1px solid #ccc !important;
                }
                
                /* Ensure mathematical formulas display correctly */
                .katex, .katex-display {
                    page-break-inside: avoid;
                }
                
                /* Ensure code blocks display correctly */
                pre, code {
                    page-break-inside: avoid;
                    white-space: pre-wrap !important;
                }
                
                /* Ensure reasonable pagination */
                h1, h2, h3, h4, h5, h6 {
                    page-break-after: avoid;
                }
                
                p {
                    page-break-inside: avoid;
                }
            </style>
            """
            
            if "<head>" in enhanced_html:
                enhanced_html = enhanced_html.replace("<head>", f"<head>{css_enhancement}", 1)
            else:
                enhanced_html = f"<head>{css_enhancement}</head>" + enhanced_html

            await page.set_content(enhanced_html, wait_until="networkidle")
            
            # Additional wait to ensure all resources are loaded
            await page.wait_for_timeout(2000)
            
            pdf_bytes = await page.pdf(
                format=req.format or "A4",
                print_background=True,
                display_header_footer=False,
                prefer_css_page_size=True,
                margin={
                    "top": req.margin_top or "10mm",
                    "right": req.margin_right or "10mm",
                    "bottom": req.margin_bottom or "10mm",
                    "left": req.margin_left or "10mm",
                }
            )
            await context.close()
            await browser.close()

            filename = (req.file_name or "document") + ".pdf"
            headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
            return Response(content=pdf_bytes, media_type="application/pdf", headers=headers)
    except Exception as e:
        logger.error(f"Failed to export PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to export PDF: {e}")


# --- Document Format Conversion ---
class ConvertRequest(BaseModel):
    target_format: str
    options: Dict[str, Any] = {}


class ConvertResponse(BaseModel):
    convert_id: str
    status: str
    message: str


@pdf_router.post("/convert/{task_id}")
async def start_convert(task_id: str, req: ConvertRequest, request: Request):
    """Start document format conversion"""
    try:
        # Get task info
        if task_id not in tasks_state:
            raise HTTPException(status_code=404, detail="Task not found")
        
        task_info = tasks_state[task_id]
        
        # Get original file path from task state
        original_file_path = task_info.get('original_file_path')
        
        # If not found, try to get from downloadable_files
        if not original_file_path and 'downloadable_files' in task_info and task_info['downloadable_files']:
            # Get the first file from downloadable_files
            first_file = next(iter(task_info['downloadable_files'].values()))
            if isinstance(first_file, dict) and 'path' in first_file:
                original_file_path = first_file['path']
            elif isinstance(first_file, str):
                original_file_path = first_file
        
        
        if not original_file_path or not os.path.exists(original_file_path):
            raise HTTPException(status_code=404, detail="Original file not found")
        
        # Import converter
        from collabtrans.converter.format_converter import converter
        
        # Get log queue for this task
        log_queue = tasks_log_queues.get(task_id)
        
        # Start conversion
        convert_id = await converter.convert(
            source_path=original_file_path,
            target_format=req.target_format,
            quality=req.options.get('quality', 'high'),
            task_id=task_id,
            log_queue=log_queue,
            options=req.options
        )
        
        return ConvertResponse(
            convert_id=convert_id,
            status="processing",
            message="Conversion started"
        )
        
    except Exception as e:
        logger.error(f"Start conversion failed: {e}")
        raise HTTPException(status_code=500, detail=f"Start conversion failed: {e}")


@pdf_router.get("/convert/{task_id}/status/{convert_id}")
async def get_convert_status(task_id: str, convert_id: str):
    """Get conversion status"""
    try:
        from collabtrans.converter.format_converter import converter
        
        status = converter.get_conversion_status(convert_id)
        
        # If conversion is completed, add download URL
        if status['status'] == 'completed':
            status['download_url'] = f"/convert/{task_id}/download/{convert_id}"
            status['filename'] = f"{Path(tasks_state.get(task_id, {}).get('original_filename', 'document')).stem}.{status.get('target_format', 'docx')}"
            
            # Log completion message
            log_queue = tasks_log_queues.get(task_id)
            if log_queue:
                try:
                    await log_queue.put("Conversion completed! File has been generated and is ready for download.")
                except Exception as e:
                    logger.warning(f"Failed to send completion log: {e}")
        
        return status
        
    except Exception as e:
        logger.error(f"Get conversion status failed: {e}")
        raise HTTPException(status_code=500, detail=f"Get conversion status failed: {e}")


@pdf_router.get("/convert/{task_id}/download/{convert_id}")
async def download_converted_file(task_id: str, convert_id: str, format: str = 'docx'):
    """Download converted file

    Args:
        format: File format to download - 'docx', 'md', or 'html'
    """
    try:
        from collabtrans.converter.format_converter import converter

        # Get converted file paths
        file_paths = converter.get_conversion_file(convert_id)
        if not file_paths:
            raise HTTPException(status_code=404, detail="Converted file not found or conversion not completed")

        # Select file based on format
        if format == 'docx':
            file_path = file_paths.get('docx')
            media_type = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        elif format == 'md':
            file_path = file_paths.get('md')
            media_type = 'text/markdown; charset=utf-8'
        elif format == 'html':
            file_path = file_paths.get('html')
            media_type = 'text/html; charset=utf-8'
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported format: {format}")

        if not file_path:
            raise HTTPException(status_code=404, detail=f"File format '{format}' not available")

        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File no longer exists")

        # Get file info
        file_stat = os.stat(file_path)
        file_size = file_stat.st_size

        # Generate filename
        original_filename = tasks_state.get(task_id, {}).get('original_filename', 'document')
        file_stem = Path(original_filename).stem
        download_filename = f"{file_stem}.{format}"

        # Return file
        return FileResponse(
            path=file_path,
            filename=download_filename,
            media_type=media_type
        )

    except Exception as e:
        logger.error(f"Download converted file failed: {e}")
        raise HTTPException(status_code=500, detail=f"Download converted file failed: {e}")


@pdf_router.get("/convert/{task_id}/formats/{convert_id}")
async def get_available_formats(task_id: str, convert_id: str):
    """Get available download formats for a conversion"""
    try:
        from collabtrans.converter.format_converter import converter

        # Get converted file paths
        file_paths = converter.get_conversion_file(convert_id)
        if not file_paths:
            return {"formats": []}

        # Get original filename for download names
        original_filename = tasks_state.get(task_id, {}).get('original_filename', 'document')
        file_stem = Path(original_filename).stem

        # Build available formats list
        formats = []
        for fmt, path in file_paths.items():
            if path and os.path.exists(path):
                file_size = os.path.getsize(path)
                formats.append({
                    "format": fmt,
                    "filename": f"{file_stem}.{fmt}",
                    "size": file_size,
                    "available": True
                })
            else:
                formats.append({
                    "format": fmt,
                    "filename": f"{file_stem}.{fmt}",
                    "size": 0,
                    "available": False
                })

        return {
            "formats": formats,
            "convert_id": convert_id,
            "task_id": task_id
        }

    except Exception as e:
        logger.error(f"Get available formats failed: {e}")
        raise HTTPException(status_code=500, detail=f"Get available formats failed: {e}")


# --- Global Configuration ---
tasks_state: Dict[str, Dict[str, Any]] = {}
tasks_log_queues: Dict[str, asyncio.Queue] = {}
tasks_log_histories: Dict[str, List[str]] = {}
MAX_LOG_HISTORY = 200
httpx_client: httpx.AsyncClient

# --- Workflow Dictionary ---
WORKFLOW_DICT: Dict[str, Type[Workflow]] = {
    "markdown_based": MarkdownBasedWorkflow,
    "txt": TXTWorkflow,
    "json": JsonWorkflow,
    "xlsx": XlsxWorkflow,
    "docx": DocxWorkflow,
    "srt": SrtWorkflow,
    "epub": EpubWorkflow,
    "html": HtmlWorkflow,
}

# --- Media Type Mapping ---
MEDIA_TYPES = {
    "html": "text/html; charset=utf-8",
    "markdown": "text/markdown; charset=utf-8",
    "markdown_zip": "application/zip",
    "txt": "text/plain; charset=utf-8",
    "json": "application/json; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv; charset=utf-8",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "srt": "text/plain; charset=utf-8",
    "epub": "application/epub+zip",
}


# --- Helper Functions ---
def _create_default_task_state() -> Dict[str, Any]:
    """Create new default task state, storing workflow instance instead of specific content"""
    return {
        "is_processing": False, "status_message": "Idle", "error_flag": False,
        "download_ready": False,
        "workflow_instance": None,  # Only used during processing
        "original_filename_stem": None, "task_start_time": 0,
        "task_end_time": 0, "current_task_ref": None,
        "original_filename": None,
        "temp_dir": None,  # Directory for storing temporary files
        "downloadable_files": {},  # Store paths and names of downloadable files
        "attachment_files": {},  # Store paths and identifiers of attachment files
    }


# --- Log Handler ---
class QueueAndHistoryHandler(logging.Handler):
    def __init__(self, queue_ref: asyncio.Queue, history_list_ref: List[str], max_history_items: int, task_id: str):
        super().__init__()
        self.queue = queue_ref
        self.history_list = history_list_ref
        self.max_history = max_history_items
        self.task_id = task_id

    def emit(self, record: logging.LogRecord):
        log_entry = self.format(record)
        print(f"[{self.task_id}] {log_entry}")
        self.history_list.append(log_entry)
        if len(self.history_list) > self.max_history:
            del self.history_list[:len(self.history_list) - self.max_history]
        if self.queue is not None:
            try:
                main_loop = getattr(app.state, "main_event_loop", None)
                if main_loop and main_loop.is_running():
                    main_loop.call_soon_threadsafe(self.queue.put_nowait, log_entry)
                else:
                    self.queue.put_nowait(log_entry)
            except asyncio.QueueFull:
                print(f"[{self.task_id}] Log queue is full. Log dropped: {log_entry}")
            except Exception as e:
                print(f"[{self.task_id}] Error putting log to queue: {e}. Log: {log_entry}")


# --- Application Lifecycle Events ---
@asynccontextmanager
async def lifespan(app: FastAPI):
    global httpx_client, AUTH_AVAILABLE
    app.state.main_event_loop = asyncio.get_running_loop()
    httpx_client = httpx.AsyncClient(timeout=httpx.Timeout(connect=5, read=120, write=300, pool=10))
    tasks_state.clear()
    tasks_log_queues.clear()
    tasks_log_histories.clear()
    global_logger.propagate = False
    # Get log level from configuration file
    from collabtrans.logger.logger import get_log_level_from_config
    global_logger.setLevel(get_log_level_from_config())
    # Use i18n logger for startup messages
    from collabtrans.logger.logger import i18n_logger
    i18n_logger.info("backend.app.startup.completed")

    # Start conversion file cleanup task
    try:
        from collabtrans.converter.format_converter import cleanup_task
        asyncio.create_task(cleanup_task())
        i18n_logger.info("backend.app.startup.cleanup_task_started")
    except Exception as e:
        i18n_logger.error("backend.app.startup.cleanup_task_failed", error=str(e))

    # Check for password recovery and reset admin password if needed
    try:
        from collabtrans.auth.password_recovery import reset_admin_password_if_recovery_enabled
        if reset_admin_password_if_recovery_enabled():
            i18n_logger.info("backend.app.startup.password_recovery_completed")
    except Exception as e:
        i18n_logger.error("backend.app.startup.password_recovery_failed", error=str(e))

    # Authentication module has been initialized at application startup
    api_url = f"http://127.0.0.1:{app.state.port_to_use}/docs"
    browser_url = f"http://127.0.0.1:{app.state.port_to_use}"
    i18n_logger.info("backend.app.startup.api_docs", url=api_url)
    i18n_logger.info("backend.app.startup.browser_access", url=browser_url)
    
    # Add API endpoints for frontend log i18n
    @app.get("/api/log-messages")
    async def get_log_messages():
        """Get log messages for frontend internationalization"""
        from collabtrans.logger.log_messages import get_frontend_log_messages
        return get_frontend_log_messages()
    
    @app.post("/api/log-language")
    async def set_log_language_endpoint(request: Request):
        """Log language is always English (simplified)"""
        # Logs are always in English, no need to change
        return {"status": "success", "language": "en", "message": "Logs are always in English"}
    
    yield
    # Clean up any remaining temporary directories
    for task_id, task_state in tasks_state.items():
        temp_dir = task_state.get("temp_dir")
        if temp_dir and os.path.isdir(temp_dir):
            try:
                shutil.rmtree(temp_dir)
                print(f"Application shutdown, cleaning up temp directory for task '{task_id}': {temp_dir}")
            except Exception as e:
                print(f"Error cleaning up temp directory for task '{task_id}' '{temp_dir}': {e}")
    await httpx_client.aclose()
    print("Application shutdown, resources cleaned up.")


# --- FastAPI Application and Route Setup ---
tags_metadata = [
    {
        "name": "Service API",
        "description": "Core service API for submitting, managing, and downloading translation tasks.",
    },
    {
        "name": "Application",
        "description": "Application-related endpoints such as metadata and default parameters.",
    },
    {
        "name": "Temp",
        "description": "Test interfaces.",
    },

]

app = FastAPI(
    docs_url=None,
    redoc_url=None,
    lifespan=lifespan,
    title="DocuTranslate API",
    description=f"""
DocuTranslate backend service API, providing document translation, status query, result download and other functions.

**Note**: All task states are stored in the service process memory, service restart will cause all task information to be lost.

### Main workflow:
1.  **`POST /service/translate`**: Submit files and translation parameters containing `workflow_type` to start a background task. The service will automatically generate and return a unique `task_id`.
2.  **`GET /service/status/{{task_id}}`**: Use the obtained `task_id` to poll this endpoint to get real-time task status.
3.  **`GET /service/logs/{{task_id}}`**: (Optional) Get real-time translation logs.
4.  **`GET /service/download/{{task_id}}/{{file_type}}`**: After task completion (when `download_ready` is `true`), download result files through this endpoint.
5.  **`GET /service/attachment/{{task_id}}/{{identifier}}`**: (Optional) If the task generates attachments (such as glossaries), download through this endpoint.
6.  **`GET /service/content/{{task_id}}/{{file_type}}`**: After task completion (when `download_ready` is `true`), get file content in JSON format.
7.  **`POST /service/cancel/{{task_id}}`**: (Optional) Cancel an ongoing task.
8.  **`POST /service/release/{{task_id}}`**: (Optional) When the task is no longer needed, release all resources it occupies on the server, including temporary files.

**Version**: {__version__}
""",
    version=__version__,
    openapi_tags=tags_metadata,
)

service_router = APIRouter(prefix="/service", tags=["Service API"])
STATIC_DIR = resource_path("static")
# Use resource_path to resolve i18n directory in both dev and PyInstaller
I18N_DIR = resource_path("i18n")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
app.mount("/i18n", StaticFiles(directory=I18N_DIR), name="i18n")

# Add favicon route to handle browser requests
@app.get("/favicon.ico")
async def favicon():
    from fastapi.responses import FileResponse
    import os
    favicon_path = os.path.join(STATIC_DIR, "favicon.ico")
    if os.path.exists(favicon_path):
        return FileResponse(favicon_path, media_type="image/x-icon")
    else:
        # Return a 204 No Content response if favicon doesn't exist
        from fastapi.responses import Response
        return Response(status_code=204)

# Initialize authentication module and add middleware and routes
if AUTH_AVAILABLE:
    try:
        # Initialize authentication module
        auth_config = AuthConfig.get_config()
        init_auth(auth_config)
        ldap_status = "enabled" if auth_config.ldap_enabled else "disabled"
        print(f"Authentication module initialized - LDAP: {ldap_status}")
        
        # Get session manager and configuration
        from collabtrans.auth import get_session_manager, get_auth_config
        session_manager = get_session_manager()
        auth_config = get_auth_config()
        
        # Add authentication middleware
        app.add_middleware(AuthMiddleware, session_manager=session_manager, config=auth_config)
        
        # Add authentication routes
        app.include_router(auth_router)
        app.include_router(auth_compat_router)
        
        print("Authentication middleware and routes added")
    except Exception as e:
        print(f"Authentication module initialization failed: {e}")
        AUTH_AVAILABLE = False


# ===================================================================
# --- Pydantic Models for Service API ---
# ===================================================================

class GlossaryAgentConfigPayload(BaseModel):
    base_url: str = Field(..., validation_alias=AliasChoices('base_url', 'baseurl'),
                          description="Base URL for the LLM API used by the Agent for glossary generation.", examples=["https://api.openai.com/v1"])
    api_key: str = Field(..., validation_alias=AliasChoices('api_key', 'key'),
                         description="LLM API key for the Agent used for glossary generation.", examples=["sk-agent-api-key"])
    api_type: Optional[str] = Field(default="openai", description="API type: 'openai' or 'ollama'. Defaults to 'openai'.", examples=["openai", "ollama"])
    model_id: str = Field(..., description="Model ID for the Agent used for glossary generation.", examples=["gpt-4-turbo"])
    to_lang: str = Field(..., description="Target language for glossary generation.", examples=["Chinese", "English"])
    temperature: float = Field(default=0.7, description="Temperature parameter for the Agent used for glossary generation.")
    max_tokens: Optional[int] = Field(default=None, description="Maximum tokens to generate. For Ollama, this maps to num_predict in options.")
    concurrent: int = Field(default=30, description="Maximum concurrent requests for the Agent.")
    timeout: int = Field(default=default_params["timeout"], description="Time to wait for API response (seconds).")
    thinking: ThinkingMode = Field(default="default", description="Thinking mode for the Agent.")
    retry: int = Field(default=default_params["retry"], description="Maximum retry count after chunk failure.")


# 1. Define base parameters shared by all workflows
class BaseWorkflowParams(BaseModel):
    skip_translate: bool = Field(default=False, description="Whether to skip translation step. If True, only document parsing and format conversion will be performed.")
    base_url: Optional[str] = Field(default=None, validation_alias=AliasChoices('base_url', 'baseurl'),
                                    description="Base URL for LLM API. Required when `skip_translate` is `False`.",
                                    examples=["https://api.openai.com/v1"])
    api_key: Optional[str] = Field(default=None, validation_alias=AliasChoices('api_key', 'key'),
                                   description="LLM API key (optional).",
                                   examples=["sk-xxxxxxxxxxxxxxxxxxxxxxxxxxxx"])
    api_type: Optional[str] = Field(default=None,
                                    description="API type: 'openai' or 'ollama'. Defaults to 'openai'.",
                                    examples=["openai", "ollama"])
    model_id: Optional[str] = Field(default=None,
                                    description="LLM model ID to use. Required when `skip_translate` is `False`.",
                                    examples=["gpt-4o"])
    to_lang: str = Field(default="Chinese", description="Target translation language.", examples=["Chinese", "English"])
    chunk_size: int = Field(default=default_params["chunk_size"], description="Chunk size for text splitting (characters).")
    concurrent: int = Field(default=default_params["concurrent"], description="Number of concurrent requests.")
    temperature: float = Field(default=default_params["temperature"], description="LLM temperature parameter.")
    max_tokens: Optional[int] = Field(default=None, description="Maximum tokens to generate. For Ollama, this maps to num_predict in options.")
    timeout: int = Field(default=default_params["timeout"], description="Time to wait for API response (seconds).")
    thinking: ThinkingMode = Field(default=default_params["thinking"], description="Thinking mode for the Agent.",
                                   examples=["default", "enable", "disable"])
    retry: int = Field(default=default_params["retry"], description="Maximum retry count after a chunk translation fails.")
    custom_prompt: Optional[str] = Field(None, description="User-defined translation prompt.", alias="custom_prompt")
    glossary_dict: Optional[Dict[str, str]] = Field(None, description="Glossary dictionary, key is original text, value is translated text.")
    glossary_generate_enable: bool = Field(default=False, description="Whether to enable automatic glossary generation.")
    glossary_agent_config: Optional[GlossaryAgentConfigPayload] = Field(None,
                                                                        description="Configuration for the Agent used for glossary generation. Required when `glossary_generate_enable` is `True`.")

    @model_validator(mode='before')
    @classmethod
    def check_translation_fields(cls, values):
        # If not skipping translation (value is False or field doesn't exist), validate that related fields must exist and not be empty
        if not values.get('skip_translate'):
            # Check for standard keys or their aliases
            if not (values.get('base_url') or values.get('baseurl')):
                raise ValueError("When `skip_translate` is `False`, `base_url` or `baseurl` field is required.")
            if not values.get('model_id'):
                raise ValueError("When `skip_translate` is `False`, `model_id` field is required.")
        # If skipping translation, no validation is performed, allowing base_url and other fields to be empty
        return values


# 2. Create independent parameter models for each workflow
class MarkdownWorkflowParams(BaseWorkflowParams):
    workflow_type: Literal['markdown_based'] = Field(..., description="Specify to use Markdown-based translation workflow.")
    convert_engine: ConvertEngineType = Field(
        "identity",
        description="Select the engine to parse files into markdown. If input file is .md, this can be `null` or not passed.",
        examples=["identity", "mineru", "docling"]
    )
    mineru_token: Optional[str] = Field(None, description="Required API token when `convert_engine` is 'mineru'.")
    formula_ocr: bool = Field(True, description="Whether to perform OCR recognition on formulas. Effective for both `mineru` and `docling`.")
    ocr_enabled: bool = Field(True, description="Whether to perform OCR recognition on images. Effective for `mineru` engine.")
    code_ocr: bool = Field(True, description="Whether to perform OCR recognition on code blocks. Only effective for `docling` engine.")
    model_version: Literal["pipeline", "vlm"] = Field("vlm",
                                                      description="Version of Mineru model, 'vlm' is the newer version. Only effective for `mineru` engine.")

    @field_validator('mineru_token')
    def check_mineru_token(cls, v, values):
        # Relaxed validation: if not provided, will be injected from local sensitive configuration on server side
        return v


class TextWorkflowParams(BaseWorkflowParams):
    workflow_type: Literal['txt'] = Field(..., description="Specify to use plain text translation workflow.")
    insert_mode: Literal["replace", "append", "prepend"] = Field(
        "replace",
        description="Insert mode for translated text. 'replace': replace original text, 'append': append after original text, 'prepend': prepend before original text."
    )
    separator: str = Field(
        "\n",
        description="Separator used to separate original text and translated text when insert_mode is 'append' or 'prepend'."
    )


class JsonWorkflowParams(BaseWorkflowParams):
    workflow_type: Literal['json'] = Field(..., description="Specify to use JSON translation workflow.")
    json_paths: List[str] = Field(
        ...,
        description="A list of jsonpath-ng expressions to specify JSON fields to be translated.",
        examples=[["$.product.name", "$.product.description", "$.features[*]"]]
    )


class XlsxWorkflowParams(BaseWorkflowParams):
    workflow_type: Literal['xlsx'] = Field(..., description="Specify to use XLSX translation workflow.")
    insert_mode: Literal["replace", "append", "prepend"] = Field(
        "replace",
        description="Insert mode for translated text. 'replace': replace original text, 'append': append after original text, 'prepend': prepend before original text."
    )
    separator: str = Field(
        "\n",
        description="Separator used to separate original text and translated text when insert_mode is 'append' or 'prepend'."
    )
    translate_regions: Optional[List[str]] = Field(
        None,
        description="Specify translation range list. Example: ['Sheet1!A1:B10', 'C:D', 'E5']. If sheet name is not specified (like 'C:D'), applies to all sheets. If None, translates all text in the entire file."
    )


class DocxWorkflowParams(BaseWorkflowParams):
    workflow_type: Literal['docx'] = Field(..., description="Specify to use DOCX translation workflow.")
    insert_mode: Literal["replace", "append", "prepend"] = Field(
        "replace",
        description="Insert mode for translated text. 'replace': replace original text, 'append': append after original text, 'prepend': prepend before original text."
    )
    separator: str = Field(
        "\n",
        description="Separator used to separate original text and translated text when insert_mode is 'append' or 'prepend'."
    )


class SrtWorkflowParams(BaseWorkflowParams):
    workflow_type: Literal['srt'] = Field(..., description="Specify to use SRT subtitle translation workflow.")
    insert_mode: Literal["replace", "append", "prepend"] = Field(
        "replace",
        description="Insert mode for translated text. 'replace': replace original text, 'append': append after original text, 'prepend': prepend before original text."
    )
    separator: str = Field(
        "\n",
        description="Separator used to separate original text and translated text when insert_mode is 'append' or 'prepend'."
    )


class EpubWorkflowParams(BaseWorkflowParams):
    workflow_type: Literal['epub'] = Field(..., description="Specify to use EPUB translation workflow.")
    insert_mode: Literal["replace", "append", "prepend"] = Field(
        "replace",
        description="Insert mode for translated text. 'replace': replace original text, 'append': append after original text, 'prepend': prepend before original text."
    )
    separator: str = Field(
        "\n",
        description="Separator used to separate original text and translated text when insert_mode is 'append' or 'prepend'."
    )


# --- HTML WORKFLOW PARAMS START ---
class HtmlWorkflowParams(BaseWorkflowParams):
    workflow_type: Literal['html'] = Field(..., description="Specify to use HTML translation workflow.")
    insert_mode: Literal["replace", "append", "prepend"] = Field(
        "replace",
        description="Insert mode for translated text. 'replace': replace original text, 'append': append after original text, 'prepend': prepend before original text."
    )
    separator: str = Field(
        " ",
        description="Separator used to separate original text and translated text when insert_mode is 'append' or 'prepend'."
    )


# --- HTML WORKFLOW PARAMS END ---


# 3. Combine them using Discriminated Union
TranslatePayload = Annotated[
    Union[
        MarkdownWorkflowParams, TextWorkflowParams, JsonWorkflowParams, XlsxWorkflowParams, DocxWorkflowParams, SrtWorkflowParams, EpubWorkflowParams, HtmlWorkflowParams],
    Field(discriminator='workflow_type')
]


# 4. Create final request body model
class TranslateServiceRequest(BaseModel):
    file_name: str = Field(..., description="Original uploaded filename with extension.",
                           examples=["my_paper.pdf", "chapter1.txt", "data.xlsx", "video.srt", "my_book.epub",
                                     "index.html"])
    file_content: str = Field(..., description="Base64 encoded file content.", examples=["JVBERi0xLjQK..."])
    payload: TranslatePayload = Field(..., description="Payload containing workflow type and corresponding parameters.")

    class Config:
        json_schema_extra = {
            "examples": [
                {
                    "file_name": "annual_report_203.pdf",
                    "file_content": "JVBERi0xLjcKJeLjz9MKMSAwIG9iago8PC9...",
                    "payload": {
                        "workflow_type": "markdown_based",
                        "skip_translate": False,
                        "base_url": "https://api.openai.com/v1",
                        "api_key": "sk-your-api-key-here",
                        "model_id": "gpt-4o",
                        "to_lang": "Chinese",
                        "chunk_size": default_params["chunk_size"],
                        "concurrent": default_params["concurrent"],
                        "temperature": default_params["temperature"],
                        "timeout": default_params["timeout"],
                        "thinking": "default",
                        "retry": default_params["retry"],
                        "glossary_generate_enable": False,
                        "convert_engine": "mineru",
                        "mineru_token": "your-mineru-token-if-any",
                        "formula_ocr": True,
                        "model_version": "vlm"
                    }
                },
                {
                    "file_name": "product_info.json",
                    "file_content": "ewogICAgImlkIjogIjEyMzQ1IiwK...",
                    "payload": {
                        "workflow_type": "json",
                        "skip_translate": False,
                        "base_url": "https://api.openai.com/v1",
                        "api_key": "sk-your-api-key-here",
                        "model_id": "gpt-4o",
                        "to_lang": "Chinese",
                        "chunk_size": default_params["chunk_size"],
                        "concurrent": default_params["concurrent"],
                        "temperature": default_params["temperature"],
                        "timeout": default_params["timeout"],
                        "thinking": "default",
                        "retry": default_params["retry"],
                        "glossary_generate_enable": False,
                        "json_paths": ["$.product.name", "$.product.description", "$.features[*]"],
                    }
                },
                {
                    "file_name": "product_list.xlsx",
                    "file_content": "UEsDBBQAAAAIA... (base64-encoded xlsx)",
                    "payload": {
                        "workflow_type": "xlsx",
                        "skip_translate": False,
                        "base_url": "https://api.openai.com/v1",
                        "api_key": "sk-your-api-key-here",
                        "model_id": "gpt-4o",
                        "to_lang": "Chinese",
                        "chunk_size": default_params["chunk_size"],
                        "concurrent": default_params["concurrent"],
                        "temperature": default_params["temperature"],
                        "timeout": default_params["timeout"],
                        "thinking": "default",
                        "retry": default_params["retry"],
                        "glossary_generate_enable": False,
                        "insert_mode": "replace",
                        "separator": "\n",
                        "translate_regions": ["Sheet1!A1:B10", "C:D"],
                        "glossary_dict": {
                            "OpenAI": "Open Artificial Intelligence",
                            "LLM": "Large Language Model"
                        }
                    }
                },
                {
                    "file_name": "complex_terms.xlsx",
                    "file_content": "UEsDBBQAAAAIA... (base64-encoded xlsx)",
                    "payload": {
                        "workflow_type": "xlsx",
                        "base_url": "https://api.openai.com/v1",
                        "api_key": "sk-your-main-translator-key",
                        "model_id": "gpt-4o",
                        "to_lang": "Chinese",
                        "retry": default_params["retry"],
                        "glossary_generate_enable": True,
                        "glossary_agent_config": {
                            "base_url": "https://api.openai.com/v1",
                            "api_key": "sk-your-agent-key-for-glossary",
                            "model_id": "gpt-4-turbo",
                            "to_lang": "Chinese",
                            "temperature": 0.7,
                            "concurrent": 30,
                            "timeout": default_params["timeout"],
                            "thinking": "default",
                            "retry": default_params["retry"]
                        }
                    }
                },
                {
                    "file_name": "contract.docx",
                    "file_content": "UEsDBBQAAAAIA... (base64-encoded docx)",
                    "payload": {
                        "workflow_type": "docx",
                        "skip_translate": False,
                        "base_url": "https://api.openai.com/v1",
                        "api_key": "sk-your-api-key-here",
                        "model_id": "gpt-4o",
                        "to_lang": "Chinese",
                        "insert_mode": "replace",
                        "separator": "\n",
                        "chunk_size": default_params["chunk_size"],
                        "concurrent": default_params["concurrent"],
                        "temperature": default_params["temperature"],
                        "timeout": default_params["timeout"],
                        "thinking": "default",
                        "retry": default_params["retry"],
                    }
                },
                {
                    "file_name": "movie.srt",
                    "file_content": "MSAKMDA6MDA6MDEsMjAwIC0tPiAwMDowMD...",
                    "payload": {
                        "workflow_type": "srt",
                        "skip_translate": False,
                        "base_url": "https://api.openai.com/v1",
                        "api_key": "sk-your-api-key-here",
                        "model_id": "gpt-4o",
                        "to_lang": "Chinese",
                        "insert_mode": "replace",
                        "separator": "\n",
                        "chunk_size": default_params["chunk_size"],
                        "concurrent": default_params["concurrent"],
                        "temperature": default_params["temperature"],
                        "timeout": default_params["timeout"],
                        "thinking": "default",
                        "retry": default_params["retry"],
                    }
                },
                {
                    "file_name": "my_book.epub",
                    "file_content": "UEsDBBQAAAAIA... (base64-encoded epub)",
                    "payload": {
                        "workflow_type": "epub",
                        "skip_translate": False,
                        "base_url": "https://api.openai.com/v1",
                        "api_key": "sk-your-api-key-here",
                        "model_id": "gpt-4o",
                        "to_lang": "Chinese",
                        "insert_mode": "replace",
                        "separator": "\n",
                        "chunk_size": default_params["chunk_size"],
                        "concurrent": default_params["concurrent"],
                        "temperature": default_params["temperature"],
                        "timeout": default_params["timeout"],
                        "thinking": "default",
                        "retry": default_params["retry"],
                    }
                },
                {
                    "file_name": "company_about_us.html",
                    "file_content": "PGh0bWw+PGhlYWQ+PHRpdGxlPkFib3V0IFVzPC90aXRsZT48L2hlYWQ+PGJvZHk+PGgxPk91ciBDb21wYW55PC9oMT48cD5XZSBhcmUgYSBsZWFkaW5nIHByb3ZpZGVyIG9mIGlubm92YXRpdmUgc29sdXRpb25zLjwvcD48L2JvZHk+PC9odG1sPg==",
                    "payload": {
                        "workflow_type": "html",
                        "skip_translate": False,
                        "base_url": "https://api.openai.com/v1",
                        "api_key": "sk-your-api-key-here",
                        "model_id": "gpt-4o",
                        "to_lang": "Chinese",
                        "insert_mode": "replace",
                        "separator": " ",
                        "chunk_size": default_params["chunk_size"],
                        "concurrent": default_params["concurrent"],
                        "temperature": default_params["temperature"],
                        "timeout": default_params["timeout"],
                        "thinking": "default",
                        "retry": default_params["retry"],
                    }
                }
            ]
        }


# --- Background Task Logic ---
async def _perform_translation(
        task_id: str,
        payload: TranslatePayload,
        file_contents: bytes,
        original_filename: str
):
    task_state = tasks_state[task_id]
    log_queue = tasks_log_queues[task_id]
    log_history = tasks_log_histories[task_id]

    task_logger = logging.getLogger(f"task.{task_id}")
    # Get log level from configuration file
    from collabtrans.logger.logger import get_log_level_from_config
    task_logger.setLevel(get_log_level_from_config())
    task_logger.propagate = False
    if task_logger.hasHandlers():
        task_logger.handlers.clear()
    task_handler = QueueAndHistoryHandler(log_queue, log_history, MAX_LOG_HISTORY, task_id=task_id)
    task_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    task_logger.addHandler(task_handler)

    task_logger.info(f"Background translation task started: file '{original_filename}', workflow: '{payload.workflow_type}'")
    task_state["status_message"] = f"Processing '{original_filename}'..."
    temp_dir = None

    try:
        # 1. Select appropriate Workflow Class based on workflow type
        workflow_class = WORKFLOW_DICT.get(payload.workflow_type)
        if not workflow_class:
            raise ValueError(f"Unsupported workflow type: '{payload.workflow_type}'")

        workflow: Workflow

        # Helper function to inject global API Key: fallback to global sensitive configuration when api_key is not provided
        def inject_global_api_key(args: dict) -> dict:
            try:
                if args.get('skip_translate'):
                    return args
                if args.get('api_key'):
                    return args

                from collabtrans.config.secrets_manager import get_secrets_manager
                from collabtrans.config.global_config import get_global_config
                secrets = get_secrets_manager()
                global_conf = get_global_config()

                base_url_raw = args.get('base_url') or ''
                base_url = base_url_raw.lower()
                logger.info(f"[DEBUG] inject_global_api_key - base_url: {base_url_raw}")
                logger.info(f"[DEBUG] inject_global_api_key - args: {args}")
                
                platform_key = None
                platform_config = None
                
                # First, try to identify platform by URL keywords
                if 'deepseek' in base_url:
                    platform_key = 'deepseek'
                elif 'openai' in base_url:
                    platform_key = 'openai'
                elif 'bigmodel' in base_url or 'zhipu' in base_url:
                    platform_key = 'zhipu'
                elif 'dashscope' in base_url or 'aliyun' in base_url:
                    platform_key = 'dashscope'
                elif 'siliconflow' in base_url:
                    platform_key = 'siliconflow'
                elif 'ark.' in base_url or 'volcengine' in base_url:
                    platform_key = 'volcengine_ark'
                elif ':11434' in base_url or 'ollama' in base_url:
                    # Ollama typically uses port 11434
                    platform_key = None  # Will try to find by URL match
                    logger.info(f"[DEBUG] inject_global_api_key - detected possible Ollama URL (port 11434)")
                
                # If platform_key identified, get config
                if platform_key:
                    platform_config = global_conf.get_ai_platform_config(platform_key)
                
                # If not identified by keywords, try to find platform by matching base_url
                if not platform_config:
                    # Clean base_url for comparison (remove trailing slashes and common paths)
                    base_url_clean = base_url_raw.strip().rstrip('/')
                    for path_prefix in ['/v1', '/v1/chat', '/v1/chat/completions', '/api', '/api/chat']:
                        if base_url_clean.endswith(path_prefix):
                            base_url_clean = base_url_clean[:-len(path_prefix)]
                    
                    # Try to find matching platform by URL
                    for key, config in global_conf.ai_platforms.items():
                        if key == 'default_platform':  # Skip special key
                            continue
                        config_url_clean = config.url.strip().rstrip('/')
                        # Remove common paths from config URL too
                        for path_prefix in ['/v1', '/v1/chat', '/v1/chat/completions', '/api', '/api/chat']:
                            if config_url_clean.endswith(path_prefix):
                                config_url_clean = config_url_clean[:-len(path_prefix)]
                        
                        # Compare URLs (case-insensitive)
                        if config_url_clean.lower() == base_url_clean.lower():
                            platform_key = key
                            platform_config = config
                            logger.info(f"[DEBUG] inject_global_api_key - matched platform by URL: {platform_key}")
                            break

                logger.info(f"[DEBUG] inject_global_api_key - detected platform_key: {platform_key}")

                # Inject api_type from platform config if not already set
                if platform_config and platform_config.api_type and not args.get('api_type'):
                    args['api_type'] = platform_config.api_type
                    logger.info(f"[DEBUG] inject_global_api_key - injected api_type: {platform_config.api_type} for platform: {platform_key}")
                elif not args.get('api_type'):
                    # If still no api_type, try to infer from URL
                    if ':11434' in base_url or 'ollama' in base_url:
                        args['api_type'] = 'ollama'
                        logger.info(f"[DEBUG] inject_global_api_key - inferred api_type: ollama from URL")

                # Only read API Key from sensitive configuration
                # For Ollama, API key is optional (local deployments don't require it)
                api_keys = secrets.get_api_keys() or {}
                key = api_keys.get(platform_key) if platform_key else None
                
                # Determine if API key is required based on api_type
                api_type = args.get('api_type', 'openai')
                is_ollama = api_type == 'ollama'
                
                if key:
                    args['api_key'] = key
                    logger.info(f"[DEBUG] inject_global_api_key - injected API key for platform: {platform_key}")
                elif is_ollama:
                    # Ollama doesn't require API key, set to None or empty string
                    args['api_key'] = None
                    logger.info(f"[DEBUG] inject_global_api_key - No API key for Ollama platform (not required for local deployments)")
                elif platform_key:
                    # For other platforms, warn if API key is missing
                    logger.warning(f"API Key for platform {platform_key} not found, please save the corresponding platform Key in the admin interface")
                # If platform_key is None, we couldn't identify the platform, but that's OK for custom/Ollama platforms
                
                # Note: We do NOT inject platform temperature here, as user's temperature setting should take priority
                # Platform temperature is only used for display purposes in the UI
                
                # Inject max_tokens from platform config if not already set and platform config exists
                if platform_config and not args.get('max_tokens') and platform_config.max_tokens:
                    args['max_tokens'] = platform_config.max_tokens
                    logger.info(f"[DEBUG] inject_global_api_key - injected max_tokens: {platform_config.max_tokens} for platform: {platform_key}")
            except Exception as e:
                logger.error(f"[DEBUG] inject_global_api_key - error: {e}")
            return args

        # Helper function: build glossary generation configuration
        def build_glossary_agent_config():
            if payload.glossary_generate_enable and payload.glossary_agent_config:
                agent_payload = payload.glossary_agent_config
                return GlossaryAgentConfig(
                    logger=task_logger,
                    **agent_payload.model_dump()
                )
            return None
        
        # Helper function: get user-selected glossary
        def get_user_glossary():
            """Get user-selected glossary"""
            try:
                # Check if username attribute exists in payload
                username = getattr(payload, 'username', None)
                if not username:
                    return {}
                from .glossary.manager import get_glossary_manager
                manager = get_glossary_manager()
                return manager.merge_user_glossaries(username)
            except Exception as e:
                logger.warning(f"Failed to get user glossary: {e}")
                return {}

        # 2. Build configuration and instantiate workflow based on payload type
        if isinstance(payload, MarkdownWorkflowParams):
            task_logger.info("Building MarkdownBasedWorkflow configuration")
            translator_args = payload.model_dump(include={
                'skip_translate', 'base_url', 'api_key', 'api_type', 'model_id', 'to_lang', 'custom_prompt',
                'temperature', 'max_tokens', 'thinking', 'chunk_size', 'concurrent', 'glossary_dict', 'timeout', 'retry'
            }, exclude_none=True)
            translator_args['glossary_generate_enable'] = payload.glossary_generate_enable
            translator_args['glossary_agent_config'] = build_glossary_agent_config()
            
            # Merge user-selected glossary
            user_glossary = get_user_glossary()
            if user_glossary:
                if 'glossary_dict' in translator_args and translator_args['glossary_dict']:
                    # Merge glossaries, user-selected glossary has higher priority
                    translator_args['glossary_dict'] = {**translator_args['glossary_dict'], **user_glossary}
                else:
                    translator_args['glossary_dict'] = user_glossary
                task_logger.info(f"User glossary loaded with {len(user_glossary)} terms")
            
            translator_args = inject_global_api_key(translator_args)
            translator_config = MDTranslatorConfig(**translator_args)

            converter_config = None
            if payload.convert_engine == 'mineru':
                # If MinerU Token is not provided by frontend, inject from local sensitive configuration
                mineru_token = payload.mineru_token
                
                # Check if token is empty or too short (normal JWT token should have 400+ characters)
                if not mineru_token or len(mineru_token) < 100:
                    try:
                        from collabtrans.config.secrets_manager import get_secrets_manager
                        sm = get_secrets_manager()
                        mineru_token = sm.get_mineru_token() or ""
                    except Exception as e:
                        task_logger.error(f"[MinerU] Failed to get token: {e}")
                        mineru_token = ""
                
                # Get mineru base URL from global configuration
                base_url = "https://mineru.net"
                try:
                    from collabtrans.config.global_config import get_global_config
                    global_conf = get_global_config()
                    # Get mineru engine configuration
                    engines = getattr(global_conf.translator_settings, 'engines', {})
                    if not isinstance(engines, dict):
                        engines = {}
                    mineru_config = engines.get("mineru", {})
                    if mineru_config.get("api_url"):
                        base_url = mineru_config.get("api_url").rstrip("/").rstrip("/api/v4")
                except Exception as e:
                    task_logger.error(f"[MinerU] Failed to get base URL: {e}")
                
                converter_config = ConverterMineruConfig(
                    logger=task_logger,
                    mineru_token=mineru_token,
                    formula_ocr=payload.formula_ocr,
                    ocr_enabled=payload.ocr_enabled,
                    model_version=payload.model_version,
                    base_url=base_url
                )
            elif payload.convert_engine == 'docling' and DOCLING_EXIST:
                # Docling remote mode has been removed, only use local mode
                converter_config = ConverterDoclingConfig(
                    logger=task_logger,
                    code_ocr=payload.code_ocr,
                    formula_ocr=payload.formula_ocr,
                    artifact=None
                )
            html_exporter_config = MD2HTMLExporterConfig(cdn=True)
            workflow_config = MarkdownBasedWorkflowConfig(
                convert_engine=payload.convert_engine, converter_config=converter_config,
                translator_config=translator_config, html_exporter_config=html_exporter_config,
                logger=task_logger
            )
            workflow = MarkdownBasedWorkflow(config=workflow_config)

        elif isinstance(payload, TextWorkflowParams):
            task_logger.info("Building TXTWorkflow configuration")
            translator_args = payload.model_dump(include={
                'skip_translate', 'base_url', 'api_key', 'api_type', 'model_id', 'to_lang', 'custom_prompt',
                'temperature', 'thinking', 'chunk_size', 'concurrent', 'glossary_dict',
                'insert_mode', 'separator', 'timeout', 'retry'
            }, exclude_none=True)
            translator_args['glossary_generate_enable'] = payload.glossary_generate_enable
            translator_args['glossary_agent_config'] = build_glossary_agent_config()
            
            # Merge user-selected glossary
            user_glossary = get_user_glossary()
            if user_glossary:
                if 'glossary_dict' in translator_args and translator_args['glossary_dict']:
                    translator_args['glossary_dict'] = {**translator_args['glossary_dict'], **user_glossary}
                else:
                    translator_args['glossary_dict'] = user_glossary
                task_logger.info(f"User glossary loaded with {len(user_glossary)} terms")
            
            translator_args = inject_global_api_key(translator_args)
            translator_config = TXTTranslatorConfig(**translator_args)

            html_exporter_config = TXT2HTMLExporterConfig(cdn=True)
            workflow_config = TXTWorkflowConfig(
                translator_config=translator_config, html_exporter_config=html_exporter_config,
                logger=task_logger
            )
            workflow = TXTWorkflow(config=workflow_config)

        elif isinstance(payload, JsonWorkflowParams):
            task_logger.info("Building JsonWorkflow configuration")
            translator_args = payload.model_dump(include={
                'skip_translate', 'base_url', 'api_key', 'api_type', 'model_id', 'to_lang', 'custom_prompt',
                'temperature', 'max_tokens', 'thinking', 'chunk_size', 'concurrent', 'glossary_dict',
                'json_paths', 'timeout', 'retry'
            }, exclude_none=True)
            translator_args['glossary_generate_enable'] = payload.glossary_generate_enable
            translator_args['glossary_agent_config'] = build_glossary_agent_config()
            
            # Merge user-selected glossary
            user_glossary = get_user_glossary()
            if user_glossary:
                if 'glossary_dict' in translator_args and translator_args['glossary_dict']:
                    translator_args['glossary_dict'] = {**translator_args['glossary_dict'], **user_glossary}
                else:
                    translator_args['glossary_dict'] = user_glossary
                task_logger.info(f"User glossary loaded with {len(user_glossary)} terms")
            
            translator_args = inject_global_api_key(translator_args)
            translator_config = JsonTranslatorConfig(**translator_args)

            html_exporter_config = Json2HTMLExporterConfig(cdn=True)
            workflow_config = JsonWorkflowConfig(
                translator_config=translator_config, html_exporter_config=html_exporter_config,
                logger=task_logger
            )
            workflow = JsonWorkflow(config=workflow_config)

        elif isinstance(payload, XlsxWorkflowParams):
            task_logger.info("Building XlsxWorkflow configuration")
            translator_args = payload.model_dump(include={
                'skip_translate', 'base_url', 'api_key', 'api_type', 'model_id', 'to_lang', 'custom_prompt',
                'temperature', 'thinking', 'chunk_size', 'concurrent',
                'insert_mode', 'separator', 'translate_regions', 'glossary_dict', 'timeout', 'retry'
            }, exclude_none=True)
            translator_args['glossary_generate_enable'] = payload.glossary_generate_enable
            translator_args['glossary_agent_config'] = build_glossary_agent_config()
            
            # Merge user-selected glossary
            user_glossary = get_user_glossary()
            if user_glossary:
                if 'glossary_dict' in translator_args and translator_args['glossary_dict']:
                    translator_args['glossary_dict'] = {**translator_args['glossary_dict'], **user_glossary}
                else:
                    translator_args['glossary_dict'] = user_glossary
                task_logger.info(f"User glossary loaded with {len(user_glossary)} terms")
            
            translator_args = inject_global_api_key(translator_args)
            translator_config = XlsxTranslatorConfig(**translator_args)

            html_exporter_config = Xlsx2HTMLExporterConfig(cdn=True)
            workflow_config = XlsxWorkflowConfig(
                translator_config=translator_config,
                html_exporter_config=html_exporter_config,
                logger=task_logger
            )
            workflow = XlsxWorkflow(config=workflow_config)

        elif isinstance(payload, DocxWorkflowParams):
            task_logger.info("Building DocxWorkflow configuration")
            # Log temperature before model_dump
            task_logger.info(f"[DEBUG] DocxWorkflowParams temperature: {payload.temperature}")
            translator_args = payload.model_dump(include={
                'skip_translate', 'base_url', 'api_key', 'api_type', 'model_id', 'to_lang', 'custom_prompt',
                'temperature', 'max_tokens', 'thinking', 'chunk_size', 'concurrent',
                'insert_mode', 'separator', 'glossary_dict', 'timeout', 'retry'
            }, exclude_none=True)
            # Log temperature after model_dump
            task_logger.info(f"[DEBUG] translator_args temperature after model_dump: {translator_args.get('temperature', 'NOT_FOUND')}")
            translator_args['glossary_generate_enable'] = payload.glossary_generate_enable
            translator_args['glossary_agent_config'] = build_glossary_agent_config()
            
            # Merge user-selected glossary
            user_glossary = get_user_glossary()
            if user_glossary:
                if 'glossary_dict' in translator_args and translator_args['glossary_dict']:
                    translator_args['glossary_dict'] = {**translator_args['glossary_dict'], **user_glossary}
                else:
                    translator_args['glossary_dict'] = user_glossary
                task_logger.info(f"User glossary loaded with {len(user_glossary)} terms")
            
            translator_args = inject_global_api_key(translator_args)
            # Log temperature after inject_global_api_key
            task_logger.info(f"[DEBUG] translator_args temperature after inject_global_api_key: {translator_args.get('temperature', 'NOT_FOUND')}")
            translator_config = DocxTranslatorConfig(**translator_args)

            html_exporter_config = Docx2HTMLExporterConfig(cdn=True)
            workflow_config = DocxWorkflowConfig(
                translator_config=translator_config,
                html_exporter_config=html_exporter_config,
                logger=task_logger
            )
            workflow = DocxWorkflow(config=workflow_config)

        elif isinstance(payload, SrtWorkflowParams):
            task_logger.info("Building SrtWorkflow configuration")
            translator_args = payload.model_dump(include={
                'skip_translate', 'base_url', 'api_key', 'api_type', 'model_id', 'to_lang', 'custom_prompt',
                'temperature', 'max_tokens', 'thinking', 'chunk_size', 'concurrent',
                'insert_mode', 'separator', 'glossary_dict', 'timeout', 'retry'
            }, exclude_none=True)
            translator_args['glossary_generate_enable'] = payload.glossary_generate_enable
            translator_args['glossary_agent_config'] = build_glossary_agent_config()
            
            # Merge user-selected glossary
            user_glossary = get_user_glossary()
            if user_glossary:
                if 'glossary_dict' in translator_args and translator_args['glossary_dict']:
                    translator_args['glossary_dict'] = {**translator_args['glossary_dict'], **user_glossary}
                else:
                    translator_args['glossary_dict'] = user_glossary
                task_logger.info(f"User glossary loaded with {len(user_glossary)} terms")
            
            translator_args = inject_global_api_key(translator_args)
            translator_config = SrtTranslatorConfig(**translator_args)

            html_exporter_config = Srt2HTMLExporterConfig(cdn=True)
            workflow_config = SrtWorkflowConfig(
                translator_config=translator_config,
                html_exporter_config=html_exporter_config,
                logger=task_logger
            )
            workflow = SrtWorkflow(config=workflow_config)

        elif isinstance(payload, EpubWorkflowParams):
            task_logger.info("Building EpubWorkflow configuration")
            translator_args = payload.model_dump(include={
                'skip_translate', 'base_url', 'api_key', 'api_type', 'model_id', 'to_lang', 'custom_prompt',
                'temperature', 'max_tokens', 'thinking', 'chunk_size', 'concurrent',
                'insert_mode', 'separator', 'glossary_dict', 'timeout', 'retry'
            }, exclude_none=True)
            translator_args['glossary_generate_enable'] = payload.glossary_generate_enable
            translator_args['glossary_agent_config'] = build_glossary_agent_config()
            
            # Merge user-selected glossary
            user_glossary = get_user_glossary()
            if user_glossary:
                if 'glossary_dict' in translator_args and translator_args['glossary_dict']:
                    translator_args['glossary_dict'] = {**translator_args['glossary_dict'], **user_glossary}
                else:
                    translator_args['glossary_dict'] = user_glossary
                task_logger.info(f"User glossary loaded with {len(user_glossary)} terms")
            
            translator_args = inject_global_api_key(translator_args)
            translator_config = EpubTranslatorConfig(**translator_args)

            html_exporter_config = Epub2HTMLExporterConfig(cdn=True)
            workflow_config = EpubWorkflowConfig(
                translator_config=translator_config,
                html_exporter_config=html_exporter_config,
                logger=task_logger
            )
            workflow = EpubWorkflow(config=workflow_config)

        # --- HTML WORKFLOW LOGIC START ---
        elif isinstance(payload, HtmlWorkflowParams):
            task_logger.info("Building HtmlWorkflow configuration")
            translator_args = payload.model_dump(include={
                'skip_translate', 'base_url', 'api_key', 'api_type', 'model_id', 'to_lang', 'custom_prompt',
                'temperature', 'max_tokens', 'thinking', 'chunk_size', 'concurrent',
                'insert_mode', 'separator', 'glossary_dict', 'timeout', 'retry'
            }, exclude_none=True)
            translator_args['glossary_generate_enable'] = payload.glossary_generate_enable
            translator_args['glossary_agent_config'] = build_glossary_agent_config()
            
            # Merge user-selected glossary
            user_glossary = get_user_glossary()
            if user_glossary:
                if 'glossary_dict' in translator_args and translator_args['glossary_dict']:
                    translator_args['glossary_dict'] = {**translator_args['glossary_dict'], **user_glossary}
                else:
                    translator_args['glossary_dict'] = user_glossary
                task_logger.info(f"User glossary loaded with {len(user_glossary)} terms")
            
            translator_args = inject_global_api_key(translator_args)
            translator_config = HtmlTranslatorConfig(**translator_args)

            workflow_config = HtmlWorkflowConfig(
                translator_config=translator_config,
                logger=task_logger
            )
            workflow = HtmlWorkflow(config=workflow_config)
        # --- HTML WORKFLOW LOGIC END ---

        else:
            raise TypeError(f"Processing logic for workflow type '{payload.workflow_type}' not implemented.")

        # 3. Read file content and execute translation
        file_stem = Path(original_filename).stem
        file_suffix = Path(original_filename).suffix
        workflow.read_bytes(content=file_contents, stem=file_stem, suffix=file_suffix)
        from collabtrans.utils.memory_utils import log_memory
        log_memory(task_logger, "app: after read_bytes", f"file size {len(file_contents) / (1024*1024):.2f} MB")
        file_contents = None  # Drop ref so only workflow's document holds bytes; reduces peak memory for large files
        await workflow.translate_async()

        # 4. Task successful, generate all downloadable files and store
        task_logger.info("Translation completed, generating temporary result files...")
        from collabtrans.utils.memory_utils import log_memory
        log_memory(task_logger, "export: before generating files", "")
        temp_dir = tempfile.mkdtemp(prefix=f"collabtrans_{task_id}_")
        task_state["temp_dir"] = temp_dir
        downloadable_files = {}
        filename_stem = task_state['original_filename_stem']

        # Check CDN availability
        is_cdn_available = True
        try:
            await httpx_client.head("https://s4.zstatic.net/ajax/libs/KaTeX/0.16.9/contrib/auto-render.min.js",
                                    timeout=3)
        except (httpx.TimeoutException, httpx.RequestError):
            is_cdn_available = False
            task_logger.warning("CDN connection failed, will use local JS for rendering.")

        # Define export function mapping
        export_map = {}

        # Fill export mapping based on workflow type
        if isinstance(workflow, HTMLExportable):
            html_config = None
            if isinstance(workflow, MarkdownBasedWorkflow):
                html_config = MD2HTMLExporterConfig(cdn=is_cdn_available)
            elif isinstance(workflow, TXTWorkflow):
                html_config = TXT2HTMLExporterConfig(cdn=is_cdn_available)
            elif isinstance(workflow, JsonWorkflow):
                html_config = Json2HTMLExporterConfig(cdn=is_cdn_available)
            elif isinstance(workflow, XlsxWorkflow):
                # For large XLSX translations, skip HTML export to avoid extra memory usage
                html_config = None
            elif isinstance(workflow, DocxWorkflow):
                # Docx translation result does not require HTML export either
                html_config = None
            elif isinstance(workflow, SrtWorkflow):
                html_config = Srt2HTMLExporterConfig(cdn=is_cdn_available)
            elif isinstance(workflow, EpubWorkflow):
                html_config = Epub2HTMLExporterConfig(cdn=is_cdn_available)

            if html_config is not None:
                export_map['html'] = (
                    lambda: workflow.export_to_html(html_config),
                    f"{filename_stem}_translated.html",
                    True,
                )
        if isinstance(workflow, MDFormatsExportable):
            export_map['markdown'] = (workflow.export_to_markdown, f"{filename_stem}_translated.md", True)
            export_map['markdown_zip'] = (workflow.export_to_markdown_zip, f"{filename_stem}_translated.zip", False)
        if isinstance(workflow, TXTExportable):
            export_map['txt'] = (workflow.export_to_txt, f"{filename_stem}_translated.txt", True)
        if isinstance(workflow, JsonExportable):
            export_map['json'] = (workflow.export_to_json, f"{filename_stem}_translated.json", True)
        if isinstance(workflow, XlsxExportable):
            export_map['xlsx'] = (workflow.export_to_xlsx, f"{filename_stem}_translated.xlsx", False)
        if isinstance(workflow, CsvExportable):
            export_map['csv'] = (workflow.export_to_csv, f"{filename_stem}_translated.csv", False)
        if isinstance(workflow, DocxExportable):
            export_map['docx'] = (workflow.export_to_docx, f"{filename_stem}_translated.docx", False)
        if isinstance(workflow, SrtExportable):
            export_map['srt'] = (workflow.export_to_srt, f"{filename_stem}_translated.srt", True)
        if isinstance(workflow, EpubExportable):
            export_map['epub'] = (workflow.export_to_epub, f"{filename_stem}_translated.epub", False)

        # Loop to generate files
        for file_type, (export_func, filename, is_string_output) in export_map.items():
            try:
                log_memory(task_logger, f"export: before {file_type}", "")
                content = await asyncio.to_thread(export_func)
                content_bytes = content.encode('utf-8') if is_string_output else content
                file_path = os.path.join(temp_dir, filename)
                with open(file_path, "wb") as f:
                    f.write(content_bytes)
                downloadable_files[file_type] = {"path": file_path, "filename": filename}
                task_logger.info(f"Successfully generated {file_type} file")
                log_memory(task_logger, f"export: after {file_type}", f"size {len(content_bytes) / (1024*1024):.2f} MB")
            except Exception as export_error:
                task_logger.error(f"Error generating {file_type} file: {export_error}", exc_info=True)

        # Process attachment files
        attachment_files = {}
        attachment_object = workflow.get_attachment()
        if attachment_object and attachment_object.attachment_dict:
            task_logger.info(f"Found {len(attachment_object.attachment_dict)} attachments, processing...")
            for identifier, doc in attachment_object.attachment_dict.items():
                try:
                    # 'doc' is a Document object
                    attachment_filename = f"{doc.stem or identifier}{doc.suffix}"
                    attachment_path = os.path.join(temp_dir, attachment_filename)
                    with open(attachment_path, "wb") as f:
                        f.write(doc.content)
                    attachment_files[identifier] = {"path": attachment_path, "filename": attachment_filename}
                    task_logger.info(f"Successfully generated attachment '{identifier}' file: {attachment_filename}")
                except Exception as attachment_error:
                    task_logger.error(f"Error generating attachment '{identifier}' file: {attachment_error}", exc_info=True)

        # 5. Task successful, update final status
        end_time = time.time()
        duration = end_time - task_state["task_start_time"]
        
        # --- Extract translation statistics (success/failure counts) ---
        import re
        translation_stats = {
            "total_chunks": 0,
            "successful_chunks": 0,
            "failed_chunks": 0
        }
        try:
            # Try to get stats from workflow.translator.agent
            _translator = getattr(workflow, 'translator', None)
            _agent_via_translator = getattr(_translator, 'agent', None) if _translator is not None else None
            if _agent_via_translator is not None:
                # Get unresolved error count and total chunks from agent
                failed_count = getattr(_agent_via_translator, 'unresolved_error_count', 0)
                task_logger.info(f"[TranslationStats] Got failed_count from agent: {failed_count}")
                # Try to get total chunks from log history
                log_history = tasks_log_histories.get(task_id, [])
                total_chunks = 0
                # First, try to get from log history
                for line in log_history:
                    # Look for "Planned to send X requests" pattern
                    match = re.search(r"Planned to send (\d+) requests", line, re.IGNORECASE)
                    if match:
                        total_chunks = int(match.group(1))
                        task_logger.info(f"[TranslationStats] Found total_chunks from log: {total_chunks}")
                        break
                
                # If not found in log history, try to get from agent's prompt count
                if total_chunks == 0 and hasattr(_agent_via_translator, 'last_prompt_count'):
                    total_chunks = getattr(_agent_via_translator, 'last_prompt_count', 0)
                    if total_chunks > 0:
                        task_logger.info(f"[TranslationStats] Found total_chunks from agent.last_prompt_count: {total_chunks}")
                
                # Also check log history for error indicators if total_chunks is 0
                if total_chunks == 0:
                    # Check for error patterns in logs
                    has_errors = False
                    for line in log_history:
                        if re.search(r"(Too many errors|All retries failed|connection error|ReadTimeout)", line, re.IGNORECASE):
                            has_errors = True
                            # Try to extract chunk count from error context
                            match = re.search(r"(\d+)\s*requests?", line, re.IGNORECASE)
                            if match:
                                total_chunks = int(match.group(1))
                            break
                    
                    # If we found errors but no chunk count, assume at least 1 chunk failed
                    if has_errors and total_chunks == 0:
                        total_chunks = 1
                        failed_count = max(failed_count, 1)  # Ensure at least 1 failed if errors detected
                
                # Update stats if we have valid data
                if total_chunks > 0:
                    translation_stats["total_chunks"] = total_chunks
                    translation_stats["failed_chunks"] = failed_count
                    translation_stats["successful_chunks"] = total_chunks - failed_count
                    task_logger.info(f"[TranslationStats] Total chunks: {total_chunks}, Successful: {translation_stats['successful_chunks']}, Failed: {failed_count}")
                elif failed_count > 0:
                    # Even if total_chunks is 0, if we have failed_count > 0, it means there were errors
                    # Set total_chunks to failed_count to indicate all chunks failed
                    translation_stats["total_chunks"] = failed_count
                    translation_stats["failed_chunks"] = failed_count
                    translation_stats["successful_chunks"] = 0
                    task_logger.info(f"[TranslationStats] Detected {failed_count} failed chunks (total_chunks was 0)")
        except Exception as _e:
            task_logger.warning(f"[TranslationStats] failed to extract stats: {_e}")
        
        # --- Attach token stats if available ---
        token_stats_obj = None
        try:
            # 1) Priority: extract from workflow.translator.agent.token_counter
            _translator = getattr(workflow, 'translator', None)
            _agent_via_translator = getattr(_translator, 'agent', None) if _translator is not None else None
            _tc_via_translator = getattr(_agent_via_translator, 'token_counter', None) if _agent_via_translator is not None else None
            if _tc_via_translator is not None:
                token_stats_obj = _tc_via_translator.get_stats()
                task_logger.info(f"[TokenStats] using translator.agent path: {token_stats_obj}")
            else:
                if _translator is None:
                    task_logger.info("[TokenStats] workflow.translator is None")
                elif _agent_via_translator is None:
                    task_logger.info("[TokenStats] workflow.translator.agent is None")
                else:
                    task_logger.info("[TokenStats] workflow.translator.agent.token_counter is None")

                # 2) Fallback: extract from workflow.agent.token_counter
                _agent = getattr(workflow, 'agent', None)
                _tc = getattr(_agent, 'token_counter', None) if _agent is not None else None
                if _tc is not None:
                    token_stats_obj = _tc.get_stats()
                    task_logger.info(f"[TokenStats] using workflow.agent path: {token_stats_obj}")
                else:
                    if _agent is None:
                        task_logger.info("[TokenStats] workflow.agent is None")
                    else:
                        task_logger.info("[TokenStats] workflow.agent.token_counter is None")
        except Exception as _e:
            task_logger.warning(f"[TokenStats] failed to extract stats: {_e}")

        # If still None, try extracting from log history (fallback)
        if token_stats_obj is None:
            try:
                log_history = tasks_log_histories.get(task_id, [])
                # Find last line containing token stats
                for line in reversed(log_history):
                    if "Token usage statistics" in line:
                        # Example: Input: 1.23K(including cached: 0.45K), Output: 2.34K(including reasoning: 0.10K), Total: 3.67K
                        def _extract(pattern):
                            m = re.search(pattern, line)
                            if not m:
                                return None
                            try:
                                v = float(m.group(1))
                                return int(v * 1000)
                            except Exception:
                                return None
                        token_stats_obj = {
                            "input_tokens": _extract(r"Input:\s*([0-9.]+)K"),
                            "cached_tokens": _extract(r"including cached:\s*([0-9.]+)K"),
                            "output_tokens": _extract(r"Output:\s*([0-9.]+)K"),
                            "reasoning_tokens": _extract(r"including reasoning:\s*([0-9.]+)K"),
                            "total_tokens": _extract(r"Total:\s*([0-9.]+)K"),
                        }
                        task_logger.info(f"[TokenStats] extracted from logs: {token_stats_obj}")
                        break
            except Exception as _e:
                task_logger.warning(f"[TokenStats] failed to extract from logs: {_e}")

        # Determine status based on translation statistics
        total_chunks = translation_stats.get("total_chunks", 0)
        failed_chunks = translation_stats.get("failed_chunks", 0)
        successful_chunks = translation_stats.get("successful_chunks", 0)
        
        # Also check log history for error indicators as fallback
        has_errors_in_logs = False
        error_reason = None
        if total_chunks == 0 and failed_chunks == 0:
            try:
                log_history = tasks_log_histories.get(task_id, [])
                for line in log_history:
                    # First, try to extract error count from "Total unresolved errors: X"
                    # Only treat as error if the count is greater than 0
                    unresolved_match = re.search(r"Total unresolved errors:\s*(\d+)", line, re.IGNORECASE)
                    if unresolved_match:
                        error_count = int(unresolved_match.group(1))
                        if error_count > 0:
                            has_errors_in_logs = True
                            failed_chunks = error_count
                            total_chunks = max(total_chunks, failed_chunks)
                            task_logger.info(f"[TranslationStats] Found errors in log: {line[:100]}")
                            task_logger.info(f"[TranslationStats] Extracted failed_chunks from log: {failed_chunks}")
                        else:
                            # "Total unresolved errors: 0" means no errors, don't treat as failure
                            task_logger.info(f"[TranslationStats] Found 'unresolved errors: 0' in log, treating as success")
                            continue
                    # Check for other error patterns (but not "unresolved errors: 0")
                    elif re.search(r"(Too many errors|All retries failed|connection error|ReadTimeout)", line, re.IGNORECASE):
                        has_errors_in_logs = True
                        task_logger.info(f"[TranslationStats] Found error in log: {line[:100]}")
                        
                        # Extract error reason
                        if re.search(r"ReadTimeout|read timeout", line, re.IGNORECASE):
                            error_reason = "ReadTimeout"
                            # Try to extract timeout value
                            timeout_match = re.search(r"timeout:\s*(\d+)s", line, re.IGNORECASE)
                            if timeout_match:
                                error_reason = f"ReadTimeout (timeout: {timeout_match.group(1)}s)"
                        elif re.search(r"ConnectTimeout|connection timeout", line, re.IGNORECASE):
                            error_reason = "ConnectionTimeout"
                        elif re.search(r"ConnectError|connection error", line, re.IGNORECASE):
                            error_reason = "ConnectionError"
                        elif re.search(r"Too many errors", line, re.IGNORECASE):
                            error_reason = "TooManyErrors"
                        break
            except Exception as e:
                task_logger.warning(f"[TranslationStats] Error checking logs: {e}")
        
        # If we have translation stats, use them to determine success/failure
        if total_chunks > 0:
            if failed_chunks == total_chunks:
                # All chunks failed
                status_message = f"Translation failed: All {total_chunks} chunks failed (took {duration:.2f} seconds)"
                error_flag = True
                download_ready = False
            elif failed_chunks > 0:
                # Partial failure
                status_message = f"Translation completed with errors: {successful_chunks}/{total_chunks} chunks succeeded, {failed_chunks} failed (took {duration:.2f} seconds)"
                error_flag = False  # Still allow download as some chunks succeeded
                download_ready = True
            else:
                # All succeeded
                status_message = f"Translation completed successfully: All {total_chunks} chunks succeeded (took {duration:.2f} seconds)"
                error_flag = False
                download_ready = True
        elif failed_chunks > 0 or has_errors_in_logs:
            # Errors detected but total_chunks is 0 - check if files were generated
            # If files were generated (even with errors), allow download
            files_generated = bool(downloadable_files)
            
            # Check if "Using original text as translation" message exists (fallback scenario)
            using_original_as_translation = False
            try:
                log_history = tasks_log_histories.get(task_id, [])
                for line in log_history:
                    if "Using original text as translation" in line or "Using original text as translation to allow file generation" in line:
                        using_original_as_translation = True
                        break
            except Exception:
                pass
            
            if failed_chunks > 0:
                base_msg = f"Translation completed with warnings: {failed_chunks} chunk(s) used original text (took {duration:.2f} seconds)" if files_generated or using_original_as_translation else f"Translation failed: {failed_chunks} chunk(s) failed (took {duration:.2f} seconds)"
            else:
                base_msg = f"Translation completed with warnings: Some segments used original text (took {duration:.2f} seconds)" if files_generated or using_original_as_translation else f"Translation failed: Errors occurred during translation (took {duration:.2f} seconds)"
            
            # Add error reason if available (but only if files were not generated)
            if error_reason and not (files_generated or using_original_as_translation):
                if error_reason == "ReadTimeout":
                    status_message = f"{base_msg} - Request timeout. The server did not respond in time. Possible causes: 1) Model is too slow, 2) Network issues, 3) Server overload. Try increasing timeout or reducing chunk_size."
                elif error_reason == "ConnectionTimeout":
                    status_message = f"{base_msg} - Connection timeout. Failed to connect to the server. Check if the server is running and accessible."
                elif error_reason == "ConnectionError":
                    status_message = f"{base_msg} - Connection error. Failed to connect to the server. Check network connectivity and server status."
                elif error_reason == "TooManyErrors":
                    status_message = f"{base_msg} - Too many errors occurred. Please check the logs for details."
                else:
                    status_message = f"{base_msg} - Error: {error_reason}"
            else:
                status_message = base_msg
            
            # If files were generated or using original as translation, allow download
            if files_generated or using_original_as_translation:
                error_flag = False  # Not a complete failure if files were generated
                download_ready = True
                task_logger.info(f"[TranslationStats] Files generated despite errors. Allowing download. files_generated={files_generated}, using_original_as_translation={using_original_as_translation}")
            else:
                error_flag = True
                download_ready = False
            
            # Update translation_stats for consistency
            if failed_chunks > 0:
                translation_stats["total_chunks"] = failed_chunks
                translation_stats["failed_chunks"] = failed_chunks
                translation_stats["successful_chunks"] = 0
            task_logger.info(f"[TranslationStats] Detected issues: failed_chunks={failed_chunks}, has_errors_in_logs={has_errors_in_logs}, error_reason={error_reason}, files_generated={files_generated}, download_ready={download_ready}")
        else:
            # No stats available, assume success (backward compatibility)
            status_message = f"Translation completed successfully in {duration:.2f} seconds"
            error_flag = False
            download_ready = True

        task_state.update({
            "status_message": status_message,
            "download_ready": download_ready,
            "error_flag": error_flag,
            "task_end_time": end_time,
            "downloadable_files": downloadable_files,
            "attachment_files": attachment_files,
            # attach token stats if the workflow exposes it via agent
            "token_stats": token_stats_obj,
            # attach translation statistics
            "translation_stats": translation_stats,
        })
        task_logger.info(f"Translation task completed, took {duration:.2f} seconds. Stats: {translation_stats}")

    except asyncio.CancelledError:
        end_time = time.time()
        duration = end_time - task_state["task_start_time"]
        task_logger.info(f"Translation task '{original_filename}' has been cancelled (took {duration:.2f} seconds).")
        task_state.update({
            "status_message": f"Translation task cancelled (took {duration:.2f} seconds)", "error_flag": False, "download_ready": False,
            "task_end_time": end_time,
        })
    except Exception as e:
        end_time = time.time()
        duration = end_time - task_state["task_start_time"]
        error_message = f"Translation failed: {e}"
        task_logger.error(error_message, exc_info=True)
        task_state.update({
            "status_message": f"Translation failed (took {duration:.2f} seconds): {e}", "error_flag": True,
            "download_ready": False,
            "task_end_time": end_time,
        })
    finally:
        # Regardless of success or failure, clean up workflow instance in memory and temporary directory (if failed)
        task_state["workflow_instance"] = None
        task_state["is_processing"] = False
        task_state["current_task_ref"] = None

        # Only clean up temp directory if task failed AND no files were generated
        # If files were generated (even with errors), keep temp directory for download
        files_generated = bool(task_state.get("downloadable_files"))
        if task_state["error_flag"] and not files_generated and temp_dir and os.path.isdir(temp_dir):
            shutil.rmtree(temp_dir)
            task_logger.info(f"Temporary directory cleaned up due to task failure (no files generated)")
            task_state["temp_dir"] = None
        elif files_generated and temp_dir:
            task_logger.info(f"Temporary directory kept for file download (files generated despite errors)")

        task_logger.info(f"Background translation task '{original_filename}' processing completed")
        task_logger.removeHandler(task_handler)


# --- Core task startup logic ---
async def _start_translation_task(
        task_id: str,
        payload: TranslatePayload,
        file_contents: bytes,
        original_filename: str
):
    if task_id not in tasks_state:
        tasks_state[task_id] = _create_default_task_state()
        tasks_log_queues[task_id] = asyncio.Queue()
        tasks_log_histories[task_id] = []
    task_state = tasks_state[task_id]

    if task_state["is_processing"] and task_state["current_task_ref"] and not task_state["current_task_ref"].done():
        raise HTTPException(status_code=429, detail=f"Task ID '{task_id}' is in progress, please try again later.")

    # If old temporary files exist, clean them up first
    if task_state.get("temp_dir") and os.path.isdir(task_state["temp_dir"]):
        shutil.rmtree(task_state["temp_dir"])

    # Create temp directory and store file
    temp_dir = tempfile.mkdtemp()
    original_file_path = os.path.join(temp_dir, original_filename)
    
    # Write file content to temp directory
    with open(original_file_path, 'wb') as f:
        f.write(file_contents)
    
    task_state.update({
        "is_processing": True,
        "status_message": "Task initializing...", "error_flag": False, "download_ready": False,
        "workflow_instance": None,
        "original_filename_stem": Path(original_filename).stem,
        "original_filename": original_filename,
        "task_start_time": time.time(), "task_end_time": 0, "current_task_ref": None,
        "temp_dir": temp_dir, "downloadable_files": {}, "attachment_files": {},
        "convert_only": payload.skip_translate,  # Use skip_translate as convert_only flag
        "original_file_path": original_file_path,  # Store original file path
    })

    log_history = tasks_log_histories[task_id]
    log_queue = tasks_log_queues[task_id]
    log_history.clear()
    while not log_queue.empty():
        try:
            log_queue.get_nowait()
        except asyncio.QueueEmpty:
            break

    initial_log_msg = f"Received new translation request: {original_filename}"
    print(f"[{task_id}] {initial_log_msg}")
    log_history.append(initial_log_msg)
    await log_queue.put(initial_log_msg)

    try:
        loop = asyncio.get_running_loop()
        task = loop.create_task(_perform_translation(task_id, payload, file_contents, original_filename))
        task_state["current_task_ref"] = task
        return {"task_started": True, "task_id": task_id, "message": "Translation task started successfully, please wait..."}
    except Exception as e:
        task_state.update({"is_processing": False, "status_message": f"Failed to start task: {e}", "error_flag": True,
                           "current_task_ref": None})
        raise HTTPException(status_code=500, detail=f"Error starting translation task: {e}")


# --- Cancel task logic ---
def _cancel_translation_logic(task_id: str):
    task_state = tasks_state.get(task_id)
    if not task_state:
        raise HTTPException(status_code=404, detail=f"Task ID '{task_id}' not found.")
    if not task_state or not task_state["is_processing"] or not task_state["current_task_ref"]:
        raise HTTPException(status_code=400, detail=f"Task ID '{task_id}' has no ongoing translation task to cancel.")

    task_to_cancel: Optional[asyncio.Task] = task_state["current_task_ref"]
    if not task_to_cancel or task_to_cancel.done():
        task_state["is_processing"] = False
        task_state["current_task_ref"] = None
        raise HTTPException(status_code=400, detail="Task has been completed or cancelled.")

    print(f"[{task_id}] Received request to cancel translation task.")
    task_to_cancel.cancel()
    task_state["status_message"] = "Cancelling task..."
    return {"cancelled": True, "message": "Cancel request sent. Please wait for status update."}


# ===================================================================
# --- Service Endpoints (/service) ---
# ===================================================================

@service_router.post(
    "/translate",
    summary="Submit translation task (unified entry point)",
    description="""
Submit translation task with file and workflow parameters. Supports two formats:

**Format 1: multipart/form-data (Recommended)**
- `file`: Binary file content (Content-Type: application/octet-stream)
- `payload`: JSON string containing workflow parameters

**Format 2: application/json (Legacy, for backward compatibility)**
- `file_name`: Original filename
- `file_content`: Base64 encoded file content
- `payload`: Workflow parameters object

- **Workflow Selection**: The `payload.workflow_type` field determines the type of this task (such as `markdown_based`, `txt`, `json`, `xlsx`, `docx`, `srt`, `epub`, `html`).
- **Dynamic Parameters**: Depending on the selected workflow, the API requires different parameter sets. Please refer to the Schema or examples below.
- **Asynchronous Processing**: This endpoint returns a task ID immediately, and the client needs to poll the status interface to get progress.
""",
    responses={
        200: {
            "description": "Translation task started successfully.",
            "content": {"application/json": {
                "example": {"task_started": True, "task_id": "a1b2c3d4", "message": "Translation task started successfully, please wait..."}}}
        },
        400: {"description": "Invalid request body, e.g., Base64 decoding failed or missing required fields."},
        429: {"description": "Server already has a task with the same ID being processed (theoretically should not happen since ID is newly generated)."},
        500: {"description": "Unknown error occurred while starting background task."},
    }
)
async def service_translate(request: Request):
    """
    Handle translation request in two formats:
    1. multipart/form-data: file + payload (JSON string)
    2. application/json: TranslateServiceRequest (legacy)
    """
    task_id = uuid.uuid4().hex[:8]
    
    # Determine request format based on content type
    content_type = request.headers.get("content-type", "").lower()
    
    if "multipart/form-data" in content_type:
        # New format: multipart/form-data
        try:
            form = await request.form()
            file = form.get("file")
            payload_str = form.get("payload")
            
            if not file:
                raise HTTPException(status_code=400, detail="File is required in multipart/form-data format")
            if not payload_str:
                raise HTTPException(status_code=400, detail="Payload is required in multipart/form-data format")
            
            # Read file content
            file_contents = await file.read()
            original_filename = file.filename or "unknown"
            
            # Parse JSON payload
            payload_dict = json.loads(payload_str)
            
            # Validate workflow_type is present
            if 'workflow_type' not in payload_dict:
                raise HTTPException(status_code=400, detail="Missing required field 'workflow_type' in payload")
            
            # Try to instantiate the payload with better error handling
            # Use TypeAdapter for better error messages with discriminated unions
            try:
                # For Pydantic v2, use TypeAdapter to validate discriminated unions
                adapter = TypeAdapter(TranslatePayload)
                payload_obj = adapter.validate_python(payload_dict)
            except ValidationError as e:
                # Pydantic validation error - provide detailed error information
                workflow_type = payload_dict.get('workflow_type', 'unknown')
                error_details = []
                for error in e.errors():
                    field_path = ' -> '.join(str(loc) for loc in error.get('loc', []))
                    error_msg = error.get('msg', 'Validation error')
                    error_type = error.get('type', 'unknown')
                    error_details.append(f"{field_path}: {error_msg} (type: {error_type})")
                
                detail_msg = f"Invalid payload for workflow_type '{workflow_type}'. "
                if error_details:
                    detail_msg += "Errors: " + "; ".join(error_details[:3])  # Show first 3 errors
                else:
                    detail_msg += str(e)
                
                raise HTTPException(status_code=400, detail=detail_msg)
            except Exception as e:
                # Handle other errors (including Union instantiation errors)
                error_msg = str(e)
                workflow_type = payload_dict.get('workflow_type', 'unknown')
                valid_types = ['markdown_based', 'txt', 'json', 'xlsx', 'docx', 'srt', 'epub', 'html']
                
                if "Union" in error_msg or "Cannot instantiate" in error_msg:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Invalid workflow_type '{workflow_type}'. Valid types are: {', '.join(valid_types)}. "
                               f"Please ensure the workflow_type matches one of the supported types and all required fields are provided."
                    )
                raise HTTPException(status_code=400, detail=f"Invalid request format: {error_msg}")
            
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail=f"Invalid JSON payload: {e}")
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid request format: {e}")
    else:
        # Legacy format: application/json
        try:
            json_data = await request.json()
            json_request = TranslateServiceRequest(**json_data)
            
            file_contents = base64.b64decode(json_request.file_content)
            original_filename = json_request.file_name
            payload_obj = json_request.payload
        except (binascii.Error, TypeError) as e:
            raise HTTPException(status_code=400, detail=f"Invalid Base64 file content: {e}")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid JSON request: {e}")

    try:
        response_data = await _start_translation_task(
            task_id=task_id,
            payload=payload_obj,
            file_contents=file_contents,
            original_filename=original_filename
        )
        return JSONResponse(content=response_data)
    except HTTPException as e:
        if e.status_code == 429:
            return JSONResponse(status_code=e.status_code, content={"task_started": False, "message": e.detail})
        if e.status_code == 500:
            return JSONResponse(status_code=e.status_code, content={"task_started": False, "message": e.detail})
        raise e


@service_router.post(
    "/cancel/{task_id}",
    summary="Cancel translation task",
    description="""Cancel an ongoing translation task based on task ID. If the task has been completed, failed, or already cancelled, an error will be returned."""
)
async def service_cancel_translate(task_id: str):
    return _cancel_translation_logic(task_id)


@service_router.post(
    "/release/{task_id}",
    summary="Release task resources",
    description="""Release all resources occupied by the task on the server based on task ID, including status, logs, and cached translation result files. If the task is in progress, it will first try to cancel the task. This operation is irreversible."""
)
async def service_release_task(task_id: str):
    if task_id not in tasks_state:
        return JSONResponse(status_code=404, content={"released": False, "message": f"Task ID '{task_id}' not found."})
    task_state = tasks_state.get(task_id)
    message_parts = []
    if task_state and task_state.get("is_processing") and task_state.get("current_task_ref"):
        try:
            print(f"[{task_id}] Task is in progress, will try to cancel before release.")
            _cancel_translation_logic(task_id)
            message_parts.append("Task has been cancelled.")
        except HTTPException as e:
            print(f"[{task_id}] Expected situation when cancelling task (may have been completed): {e.detail}")
            message_parts.append(f"Task cancellation step skipped (may have been completed or cancelled).")

    if task_state:
        temp_dir = task_state.get("temp_dir")
        if temp_dir and os.path.isdir(temp_dir):
            try:
                shutil.rmtree(temp_dir)
                message_parts.append("Temporary files cleaned up.")
                print(f"[{task_id}] Temporary directory '{temp_dir}' has been deleted.")
            except Exception as e:
                message_parts.append(f"Error cleaning up temporary files: {e}.")
                print(f"[{task_id}] Error deleting temporary directory '{temp_dir}': {e}")

    tasks_state.pop(task_id, None)
    tasks_log_queues.pop(task_id, None)
    tasks_log_histories.pop(task_id, None)
    print(f"[{task_id}] Resources successfully released.")
    message_parts.append(f"Resources for task '{task_id}' have been released.")
    return JSONResponse(content={"released": True, "message": " ".join(message_parts)})


@service_router.get(
    "/status/{task_id}",
    summary="Get task status",
    description="Get the current status of a task based on task ID. When `download_ready` is `true`, the `downloads` and `attachment` objects will contain available download links.",
    responses={
        200: {
            "description": "Successfully retrieved task status.",
            "content": {
                "application/json": {
                    "examples": {
                        "processing": {
                            "summary": "In Progress",
                            "value": {
                                "task_id": "a1b2c3d4", "is_processing": True,
                                "status_message": "Processing 'annual_report.pdf'...",
                                "error_flag": False, "download_ready": False, "original_filename_stem": "annual_report",
                                "original_filename": "annual_report.pdf", "task_start_time": 1678889400.0,
                                "task_end_time": 0, "downloads": {}, "attachment": {}
                            }
                        },
                        "completed_markdown": {
                            "summary": "Completed (Markdown)",
                            "value": {
                                "task_id": "b2865b93", "is_processing": False,
                                "status_message": "Translation completed successfully in 123.45 seconds",
                                "error_flag": False, "download_ready": True, "original_filename_stem": "my_paper",
                                "original_filename": "my_paper.pdf", "task_start_time": 1678889400.123,
                                "task_end_time": 1678889523.573,
                                "downloads": {
                                    "html": "/service/download/b2865b93/html",
                                    "markdown": "/service/download/b2865b93/markdown",
                                    "markdown_zip": "/service/download/b2865b93/markdown_zip"
                                },
                                "attachment": {}
                            }
                        },
                        "completed_with_attachment": {
                            "summary": "Completed (with attachments)",
                            "value": {
                                "task_id": "g1h2i3j4", "is_processing": False,
                                "status_message": "Translation completed successfully in 125.00 seconds",
                                "error_flag": False, "download_ready": True,
                                "original_filename_stem": "complex_document",
                                "original_filename": "complex_document.docx",
                                "task_start_time": 1678891000.0,
                                "task_end_time": 1678891125.0,
                                "downloads": {
                                    "docx": "/service/download/g1h2i3j4/docx",
                                    "html": "/service/download/g1h2i3j4/html"
                                },
                                "attachment": {
                                    "glossary": "/service/attachment/g1h2i3j4/glossary"
                                }
                            }
                        },
                        "completed_xlsx": {
                            "summary": "Completed (XLSX)",
                            "value": {
                                "task_id": "d7e8f9a0",
                                "is_processing": False,
                                "status_message": "Translation completed successfully in 18.99 seconds",
                                "error_flag": False,
                                "download_ready": True,
                                "original_filename_stem": "sales_data",
                                "original_filename": "sales_data.xlsx",
                                "task_start_time": 1678889600.0,
                                "task_end_time": 1678889618.99,
                                "downloads": {
                                    "xlsx": "/service/download/d7e8f9a0/xlsx",
                                    "csv": "/service/download/d7e8f9a0/csv",
                                    "html": "/service/download/d7e8f9a0/html"
                                },
                                "attachment": {}
                            }
                        },
                        "completed_docx": {
                            "summary": "Completed (DOCX)",
                            "value": {
                                "task_id": "f8a9c1b2", "is_processing": False,
                                "status_message": "Translation completed successfully in 25.10 seconds",
                                "error_flag": False, "download_ready": True, "original_filename_stem": "contract",
                                "original_filename": "contract.docx", "task_start_time": 1678889500.123,
                                "task_end_time": 1678889525.223,
                                "downloads": {
                                    "docx": "/service/download/f8a9c1b2/docx",
                                    "html": "/service/download/f8a9c1b2/html"
                                },
                                "attachment": {}
                            }
                        },
                        "completed_epub": {
                            "summary": "Completed (EPUB)",
                            "value": {
                                "task_id": "e9b8d7c6", "is_processing": False,
                                "status_message": "Translation completed successfully in 45.32 seconds",
                                "error_flag": False, "download_ready": True, "original_filename_stem": "my_book",
                                "original_filename": "my_book.epub", "task_start_time": 1678890000.0,
                                "task_end_time": 1678890045.32,
                                "downloads": {
                                    "epub": "/service/download/e9b8d7c6/epub",
                                    "html": "/service/download/e9b8d7c6/html"
                                },
                                "attachment": {}
                            }
                        },
                        # --- HTML STATUS EXAMPLE START ---
                        "completed_html": {
                            "summary": "Completed (HTML)",
                            "value": {
                                "task_id": "a1b2c3d4", "is_processing": False,
                                "status_message": "Translation completed successfully in 15.78 seconds",
                                "error_flag": False, "download_ready": True, "original_filename_stem": "about_us",
                                "original_filename": "about_us.html", "task_start_time": 1678890100.0,
                                "task_end_time": 1678890115.78,
                                "downloads": {
                                    "html": "/service/download/a1b2c3d4/html"
                                },
                                "attachment": {}
                            }
                        },
                        # --- HTML STATUS EXAMPLE END ---
                        "error": {
                            "summary": "Failed",
                            "value": {
                                "task_id": "c3d4e5f6", "is_processing": False,
                                "status_message": "Translation failed: LLM API key is invalid",
                                "error_flag": True, "download_ready": False, "original_filename_stem": "bad_config",
                                "original_filename": "bad_config.json", "task_start_time": 1678889600.0,
                                "task_end_time": 1678889610.0, "downloads": {}, "attachment": {}
                            }
                        }
                    }
                }
            }
        },
        404: {"description": "Specified task ID does not exist."},
    }
)
async def service_get_status(
        task_id: str = FastApiPath(..., description="ID of the task to query status for", examples=["b2865b93"])):
    task_state = tasks_state.get(task_id)
    if not task_state:
        raise HTTPException(status_code=404, detail=f"Task ID '{task_id}' not found.")

    downloads = {}
    if task_state.get("download_ready") and task_state.get("downloadable_files"):
        for file_type in task_state["downloadable_files"].keys():
            downloads[file_type] = f"/service/download/{task_id}/{file_type}"

    attachments = {}
    if task_state.get("download_ready") and task_state.get("attachment_files"):
        for identifier in task_state["attachment_files"].keys():
            attachments[identifier] = f"/service/attachment/{task_id}/{identifier}"

    return JSONResponse(content={
        "task_id": task_id,
        "is_processing": task_state["is_processing"],
        "status_message": task_state["status_message"],
        "error_flag": task_state["error_flag"],
        "download_ready": task_state["download_ready"],
        "original_filename_stem": task_state["original_filename_stem"],
        "original_filename": task_state.get("original_filename"),
        "task_start_time": task_state["task_start_time"],
        "task_end_time": task_state["task_end_time"],
        "downloads": downloads,
        "attachment": attachments,
        # expose token stats if available
        "token_stats": task_state.get("token_stats")
    })


@service_router.get(
    "/logs/{task_id}",
    summary="Get task incremental logs",
    description="""Get task incremental logs in streaming mode. Each time the client calls this interface, it returns new log lines generated since the last call. This is very useful for real-time display of translation progress. If the task ID does not exist, 404 is returned."""
)
async def service_get_logs(task_id: str):
    if task_id not in tasks_log_queues:
        raise HTTPException(status_code=404, detail=f"Log queue for task ID '{task_id}' not found.")
    log_queue = tasks_log_queues[task_id]
    new_logs = []
    while not log_queue.empty():
        try:
            new_logs.append(log_queue.get_nowait())
            log_queue.task_done()
        except asyncio.QueueEmpty:
            break
    return JSONResponse(content={"logs": new_logs})


FileType = Literal["markdown", "markdown_zip", "html", "txt", "json", "xlsx", "csv", "docx", "srt", "epub"]


@service_router.get(
    "/download/{task_id}/{file_type}",
    summary="Download translation result files",
    responses={
        200: {
            "description": "Successfully returned file stream. Filename is specified via Content-Disposition header. Content-Type is application/octet-stream.",
            "content": {
                "application/octet-stream": {"schema": {"type": "string", "format": "binary"}},
            }
        },
        404: {"description": "Task ID does not exist, or the task does not support the requested file type, or temporary files have been lost."},
        500: {"description": "Internal error occurred while reading file on server."}
    }
)
async def service_download_file(
        task_id: str = FastApiPath(..., description="ID of completed task", examples=["b2865b93"]),
        file_type: FileType = FastApiPath(..., description="File type to download.",
                                          examples=["html", "json", "csv", "docx", "srt", "epub"])
):
    task_state = tasks_state.get(task_id)
    if not task_state:
        raise HTTPException(status_code=404, detail=f"Task ID '{task_id}' not found.")

    file_info = task_state.get("downloadable_files", {}).get(file_type)
    if not file_info or not os.path.exists(file_info.get("path")):
        raise HTTPException(status_code=404,
                            detail=f"Task '{task_id}' does not support downloading '{file_type}' type files, or files have been lost.")

    file_path = file_info["path"]
    filename = file_info["filename"]
    # Use application/octet-stream for all file downloads
    media_type = "application/octet-stream"

    return FileResponse(path=file_path, media_type=media_type, filename=filename)


@service_router.get(
    "/attachment/{task_id}/{identifier}",
    summary="Download attachment files",
    description="Download additional files generated during translation based on task ID and attachment identifier, such as automatically generated glossaries.",
    responses={
        200: {
            "description": "Successfully returned file stream. Filename is specified via Content-Disposition header.",
            "content": {
                "application/octet-stream": {"schema": {"type": "string", "format": "binary"}},
            }
        },
        404: {"description": "Task ID does not exist, or the task has no specified attachment, or temporary files have been lost."},
    }
)
async def service_download_attachment(
        task_id: str = FastApiPath(..., description="ID of completed task", examples=["g1h2i3j4"]),
        identifier: str = FastApiPath(..., description="Identifier of the attachment to download.", examples=["glossary"])
):
    task_state = tasks_state.get(task_id)
    if not task_state:
        raise HTTPException(status_code=404, detail=f"Task ID '{task_id}' not found.")

    attachment_info = task_state.get("attachment_files", {}).get(identifier)
    if not attachment_info or not os.path.exists(attachment_info.get("path")):
        raise HTTPException(status_code=404,
                            detail=f"Task '{task_id}' has no attachment with identifier '{identifier}', or files have been lost.")

    file_path = attachment_info["path"]
    filename = attachment_info["filename"]

    # Use a generic media type as attachments can be of various formats
    media_type = "application/octet-stream"

    return FileResponse(path=file_path, media_type=media_type, filename=filename)


@service_router.get(
    "/content/{task_id}/{file_type}",
    summary="Download translation result content (JSON)",
    description="""
Get content of specified file type in JSON format instead of downloading files directly.

- **Return Structure**: Returns a JSON object containing filename, file type, and Base64-encoded string of file content.
- **Content Encoding**: File content is always **Base64** encoded, and the client needs to decode it to use.
""",
    responses={
        200: {
            "description": "Successfully returned file content.",
            "content": {"application/json": {"examples": {
                "html_base64": {
                    "summary": "HTML Content (Base64)",
                    "value": {
                        "file_type": "html",
                        "filename": "my_doc_translated.html",
                        "content": "PGh0bWw+PGhlYWQ+..."
                    }
                },
                "docx_base64": {
                    "summary": "DOCX Content (Base64)",
                    "value": {
                        "file_type": "docx",
                        "filename": "my_doc_translated.docx",
                        "content": "UEsDBBQAAAAIA... (base64-encoded string)"
                    }
                },
                "epub_base64": {
                    "summary": "EPUB Content (Base64)",
                    "value": {
                        "file_type": "epub",
                        "filename": "my_book_translated.epub",
                        "content": "UEsDBBQAAAAIA... (base64-encoded string)"
                    }
                }
            }}}
        },
        404: {"description": "Task ID does not exist, or the task does not support the requested file type, or temporary files have been lost."},
        500: {"description": "Internal error occurred while reading file on server."}
    }
)
async def service_content(
        task_id: str = FastApiPath(..., description="ID of completed task", examples=["b2865b93"]),
        file_type: FileType = FastApiPath(..., description="File type to get content for.",
                                          examples=["html", "json", "csv", "docx", "srt", "epub"])
):
    task_state = tasks_state.get(task_id)
    if not task_state:
        raise HTTPException(status_code=404, detail=f"Task ID '{task_id}' not found.")

    file_info = task_state.get("downloadable_files", {}).get(file_type)
    if not file_info or not os.path.exists(file_info.get("path")):
        raise HTTPException(status_code=404,
                            detail=f"Task '{task_id}' does not support getting '{file_type}' type content, or files have been lost.")

    file_path = file_info["path"]
    filename = file_info["filename"]

    try:
        with open(file_path, "rb") as f:
            content_bytes = f.read()
        final_content = base64.b64encode(content_bytes).decode('utf-8')
        return JSONResponse(content={
            "file_type": file_type,
            "filename": filename,
            "content": final_content
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal error occurred while reading file: {e}")


# ===================================================================
# --- Application main routes and startup ---
# ===================================================================
@service_router.get("/engin-list", tags=["Application"], description="Return available conversion engines")
async def service_get_engin_list():
    engin_list = ["mineru"]
    if DOCLING_EXIST: engin_list.append("docling")
    return JSONResponse(content=engin_list)


@service_router.get("/task-list", tags=["Application"], description="Return list of ongoing task_ids")
async def service_get_task_list(): return JSONResponse(content=list(tasks_state.keys()))


@service_router.get("/default-params", tags=["Application"], description="Return some default parameters")
def service_get_default_params(): return JSONResponse(content=default_params)


@service_router.get("/meta", tags=["Application"], description="Return software version number")
async def service_get_app_version(): return JSONResponse(content={"version": __version__})


def _get_redirect_url(request: Request, path: str) -> str:
    """Build redirect URL with root_path prefix for reverse proxy deployment."""
    root_path = request.scope.get("root_path", "") or ""
    return root_path + path


@app.get("/", response_class=HTMLResponse, include_in_schema=False)
async def main_page(request: Request):
    # Redirect to login page if not authenticated
    try:
        from collabtrans.auth import get_session_manager
        session_manager = get_session_manager()
        if not await session_manager.is_authenticated(request):
            return RedirectResponse(url=_get_redirect_url(request, "/login?next=/"), status_code=302)
    except Exception:
        # Continue directly when authentication module is unavailable
        pass

    index_path = Path(STATIC_DIR) / "index.html"
    if not index_path.exists(): raise HTTPException(status_code=404, detail="index.html not found")
    no_cache_headers = {"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0", "Pragma": "no-cache",
                        "Expires": "0"}
    return FileResponse(index_path, headers=no_cache_headers)


@app.get("/settings", response_class=HTMLResponse, include_in_schema=False)
async def settings_page(request: Request):
    # Redirect to login page if not authenticated
    try:
        from collabtrans.auth import get_session_manager
        session_manager = get_session_manager()
        if not await session_manager.is_authenticated(request):
            return RedirectResponse(url=_get_redirect_url(request, "/login?next=/settings"), status_code=302)
        # Allow super admin, admin group members, or app admin (glossary management role) to access
        try:
            user = await session_manager.get_user(request)
            # app admin: ldap_glossary / ldap_app roles are allowed
            can_app_admin = False
            try:
                role_val = getattr(user, 'role', None).value if getattr(user, 'role', None) is not None else ''
                role_val_lc = role_val.lower()
                can_app_admin = (role_val_lc == 'ldap_glossary' or role_val_lc == 'ldap_app') or (hasattr(user, 'can_access_glossary_management') and user.can_access_glossary_management())
            except Exception:
                can_app_admin = False
            if not user or not (user.is_super_admin() or user.is_admin() or can_app_admin):
                return RedirectResponse(url=_get_redirect_url(request, "/"), status_code=302)
        except Exception:
            return RedirectResponse(url=_get_redirect_url(request, "/"), status_code=302)
    except Exception:
        # Continue directly when authentication module is unavailable
        pass

    index_path = Path(STATIC_DIR) / "settings.html"
    if not index_path.exists(): raise HTTPException(status_code=404, detail="settings.html not found")
    no_cache_headers = {"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0", "Pragma": "no-cache",
                        "Expires": "0"}
    return FileResponse(index_path, headers=no_cache_headers)

@app.get("/admin", response_class=HTMLResponse, include_in_schema=False)
async def main_page_admin(request: Request):
    # Redirect to login page if not authenticated
    try:
        from collabtrans.auth import get_session_manager
        session_manager = get_session_manager()
        if not await session_manager.is_authenticated(request):
            return RedirectResponse(url=_get_redirect_url(request, "/login?next=/admin"), status_code=302)
    except Exception:
        pass

    index_path = Path(STATIC_DIR) / "index.html"
    if not index_path.exists(): raise HTTPException(status_code=404, detail="index.html not found")
    no_cache_headers = {"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0", "Pragma": "no-cache",
                        "Expires": "0"}
    return FileResponse(index_path, headers=no_cache_headers)


@app.get("/docs", include_in_schema=False)
async def custom_swagger_ui_html():
    return get_swagger_ui_html(
        openapi_url=app.openapi_url,
        title=app.title + " - Swagger UI",
        oauth2_redirect_url=app.swagger_ui_oauth2_redirect_url,
        swagger_js_url="/static/swagger/swagger.js",
        swagger_css_url="/static/swagger/swagger.css",
    )


@app.get(app.swagger_ui_oauth2_redirect_url, include_in_schema=False)
async def swagger_ui_redirect():
    return get_swagger_ui_oauth2_redirect_html()


@app.middleware("http")
async def https_redirect_middleware(request: Request, call_next):
    try:
        from collabtrans.config.local_config import LocalConfig
        local_config = LocalConfig.load_from_file()
        if local_config.https.enabled and local_config.https.force_redirect:
            proto = request.headers.get('x-forwarded-proto') or request.url.scheme
            host = request.headers.get('host')
            if proto == 'http' and host:
                https_url = str(request.url).replace('http://', 'https://', 1)
                return RedirectResponse(url=https_url, status_code=308)
    except Exception:
        pass
    return await call_next(request)

@app.get("/redoc", include_in_schema=False)
async def redoc_html():
    return get_redoc_html(
        openapi_url=app.openapi_url,
        title=app.title + " - ReDoc",
        redoc_js_url="/static/redoc/redoc.js",
    )


@app.post("/temp/translate", tags=["Temp"])
async def temp_translate(
        base_url: str = Body(...), api_key: str = Body(...), model_id: str = Body(...),
        mineru_token: Optional[str] = Body(None), file_name: str = Body(...), file_content: str = Body(...),
        to_lang: str = Body("Chinese"), concurrent: int = Body(default_params["concurrent"]),
        temperature: float = Body(default_params["temperature"]),
        thinking: ThinkingMode = Body(default_params["thinking"]),
        chunk_size: int = Body(default_params["chunk_size"]), custom_prompt: Optional[str] = Body(None),
        model_version: Literal["pipeline", "vlm"] = Body("vlm"),
        glossary_dict: Optional[Dict[str, str]] = Body(None),
):
    file_name = Path(file_name)
    try:
        decoded_content = base64.b64decode(file_content)
    except (ValueError, binascii.Error):
        decoded_content = file_content.encode('utf-8')
    try:
        workflow_config = MarkdownBasedWorkflowConfig(
            convert_engine="mineru",
            converter_config=ConverterMineruConfig(mineru_token=mineru_token, model_version=model_version),
            translator_config=MDTranslatorConfig(base_url=base_url, api_key=api_key, model_id=model_id,
                                                 to_lang=to_lang, custom_prompt=custom_prompt, temperature=temperature,
                                                 thinking=thinking, chunk_size=chunk_size, concurrent=concurrent,
                                                 glossary_dict=glossary_dict),
            html_exporter_config=MD2HTMLExporterConfig()
        )
        workflow = MarkdownBasedWorkflow(workflow_config)
        workflow.read_bytes(content=decoded_content, stem=file_name.stem, suffix=file_name.suffix)
        await workflow.translate_async()
        return {"success": True, "content": workflow.export_to_markdown()}
    except Exception as e:
        global_logger.error(f"Temporary translation interface error: {e.__repr__()}", exc_info=True)
        return {"success": False, "reason": e.__repr__()}


app.include_router(service_router)
app.include_router(pdf_router)

# ---------------------------
# Preview endpoints (DOCX/XLSX -> HTML)
# ---------------------------

@preview_router.post("/preview/docx", response_class=HTMLResponse)
async def preview_docx(file: UploadFile = File(...)):
    """Convert a DOCX file to HTML for preview using mammoth."""
    try:
        import mammoth  # type: ignore
    except Exception:
        raise HTTPException(status_code=503, detail="DOCX preview requires 'mammoth' package. Please install it.")

    try:
        from io import BytesIO
        content = await file.read()
        # mammoth expects a file-like object with .seek
        result = mammoth.convert_to_html(BytesIO(content))
        html = result.value or ""
        wrapped = (
            "<html><head><meta charset='utf-8'>"
            "<style>body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,'Helvetica Neue',Arial;padding:12px}</style>"
            "</head><body>" + html + "</body></html>"
        )
        return HTMLResponse(content=wrapped)
    except Exception as e:
        logger.exception("DOCX preview failed: %s", e)
        raise HTTPException(status_code=500, detail=f"DOCX preview failed: {e}")


@preview_router.post("/preview/xlsx", response_class=HTMLResponse)
async def preview_xlsx(file: UploadFile = File(...), max_rows: int = 200):
    """Convert an XLSX file to HTML for preview. Prefer pandas+openpyxl; fallback to openpyxl only."""
    use_pandas = False
    try:
        import pandas as pd  # type: ignore
        use_pandas = True
    except Exception:
        pass

    try:
        data = await file.read()
        tabs_html: list[str] = []
        if use_pandas:
            from io import BytesIO
            import pandas as pd  # type: ignore
            xls = pd.ExcelFile(BytesIO(data))
            for sheet in xls.sheet_names:
                df = xls.parse(sheet)
                if max_rows and len(df) > max_rows:
                    df = df.head(max_rows)
                table_html = df.to_html(border=0, classes="table table-sm table-striped", index=False)
                tabs_html.append(f"<div class='sheet'><h4>{sheet}</h4>{table_html}</div>")
        else:
            try:
                from openpyxl import load_workbook  # type: ignore
            except Exception:
                raise HTTPException(status_code=503, detail="XLSX preview requires 'pandas' or 'openpyxl'.")
            from io import BytesIO
            wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
            for ws in wb.worksheets:
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if max_rows and i >= max_rows:
                        break
                    rows.append([(c if c is not None else "") for c in row])
                thead = ""
                tbody = ""
                if rows:
                    thead = "<thead><tr>" + "".join(f"<th>{str(v)}</th>" for v in rows[0]) + "</tr></thead>"
                    for r in rows[1:]:
                        tbody += "<tr>" + "".join(f"<td>{str(v)}</td>" for v in r) + "</tr>"
                table_html = f"<table class='table table-sm table-striped'>{thead}<tbody>{tbody}</tbody></table>"
                tabs_html.append(f"<div class='sheet'><h4>{ws.title}</h4>{table_html}</div>")

        wrapped = (
            "<html><head><meta charset='utf-8'>"
            "<link rel=\"stylesheet\" href=\"https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css\">"
            "<style>body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,'Helvetica Neue',Arial;padding:12px}.sheet{margin-bottom:24px}</style>"
            "</head><body>" + "\n".join(tabs_html) + "</body></html>"
        )
        return HTMLResponse(content=wrapped)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("XLSX preview failed: %s", e)
        raise HTTPException(status_code=500, detail=f"XLSX preview failed: {e}")

app.include_router(preview_router)


def find_free_port(start_port):
    port = start_port
    while True:
        with closing(socket.socket(socket.AF_INET, socket.SOCK_STREAM)) as sock:
            if sock.connect_ex(('127.0.0.1', port)) != 0: return port
            port += 1


def run_app(port: int | None = None):
    # Automatically create local_secrets.json file on first deployment
    # Configuration file priority:
    # Windows/Linux override:
    # 0) COLLABTRANS_CONFIG_PATH env dir if set (Windows default: C:\\Users\\Public\\collabtrans)
    # 1) Environment-based path (production: /etc/collabtrans/, development: project root)
    # Common:
    # 2) local_secrets.json in executable directory (packaged configuration, legacy fallback)
    # 3) local_secrets.json in current directory (development environment, legacy fallback)

    # 0) Environment-configured directory
    env_dir = os.environ.get("COLLABTRANS_CONFIG_PATH")
    # Windows default runtime configuration directory
    if not env_dir and os.name == "nt":
        env_dir = r"C:\\Users\\Public\\collabtrans"
    secrets_path = None
    if env_dir:
        env_secrets = os.path.join(env_dir, "local_secrets.json")
        if os.path.exists(env_secrets):
            secrets_path = env_secrets
            logger.debug(f"Using env secrets config: {secrets_path}")

    # 1) Environment-based path (production or development)
    if secrets_path is None:
        if is_production():
            # Production: use /etc/collabtrans/
            prod_secrets = get_prod_config_path("local_secrets.json")
            if prod_secrets.exists():
                secrets_path = str(prod_secrets)
                print(f"Using production secrets config: {secrets_path}")
        else:
            # Development: use project root
            dev_secrets = get_dev_config_path("local_secrets.json")
            if dev_secrets.exists():
                secrets_path = str(dev_secrets)
                print(f"Using development secrets config: {secrets_path}")

    # 2) Legacy fallback: System directory (non-Windows)
    if secrets_path is None:
        system_secrets_path = "/etc/collabtrans/local_secrets.json"
        system_dir_exists = os.path.exists("/etc/collabtrans")
        if system_dir_exists and os.path.exists(system_secrets_path):
            secrets_path = system_secrets_path
            print(f"Using system secrets config (legacy): {secrets_path}")

    # 3) Legacy fallback: Executable directory or current directory
    if secrets_path is None:
        import sys
        if getattr(sys, 'frozen', False):
            # PyInstaller packaged environment
            exe_dir = os.path.dirname(sys.executable)
            exe_secrets_path = os.path.join(exe_dir, "local_secrets.json")
            if os.path.exists(exe_secrets_path):
                secrets_path = exe_secrets_path
                print(f"Using executable directory secrets config (legacy): {secrets_path}")
            else:
                secrets_path = os.path.join(os.getcwd(), "local_secrets.json")
                print(f"Using working directory secrets config (legacy): {secrets_path}")
        else:
            # Development environment
            secrets_path = os.path.join(os.getcwd(), "local_secrets.json")
            print(f"Using working directory secrets config (legacy): {secrets_path}")
    
    # Check if configuration file needs to be created
    if not os.path.exists(secrets_path):
        # Determine template file path based on environment
        template_path = None
        
        # 0) Environment-configured directory
        if env_dir:
            env_template = os.path.join(env_dir, "local_secrets.json.template")
            if os.path.exists(env_template):
                template_path = env_template
        
        # 1) Environment-based path (production or development)
        if template_path is None:
            if is_production():
                prod_template = get_prod_config_path("local_secrets.json.template")
                if prod_template.exists():
                    template_path = str(prod_template)
            else:
                dev_template = get_dev_config_path("local_secrets.json.template")
                if dev_template.exists():
                    template_path = str(dev_template)
        
        # 2) Legacy fallback: System directory
        if template_path is None:
            system_template_path = "/etc/collabtrans/local_secrets.json.template"
            system_dir_exists = os.path.exists("/etc/collabtrans")
            if system_dir_exists and os.path.exists(system_template_path):
                template_path = system_template_path
        
        # 3) Legacy fallback: Executable directory or current directory
        if template_path is None:
            import sys
            exe_template_path = os.path.join(os.path.dirname(sys.executable), "local_secrets.json.template") if getattr(sys, 'frozen', False) else None
            local_template_path = os.path.join(os.getcwd(), "local_secrets.json.template")
            if exe_template_path and os.path.exists(exe_template_path):
                template_path = exe_template_path
            elif os.path.exists(local_template_path):
                template_path = local_template_path
        
        if template_path:
            try:
                import shutil
                shutil.copy2(template_path, secrets_path)
                print(f"First deployment: Automatically created {secrets_path} from template")
                print("Please edit this file to set your API keys and admin password")
            except Exception as e:
                print(f"Failed to automatically create {secrets_path}: {e}")
        else:
            print("Warning: No local_secrets.json.template found for first deployment setup")
    
    initial_port = port or int(os.environ.get("DOCUTRANSLATE_PORT", 8020))
    try:
        port_to_use = find_free_port(initial_port)
        if port_to_use != initial_port: print(f"Port {initial_port} is occupied, using port {port_to_use} instead")
        print(f"Starting DocuTranslate WebUI version: {__version__}")
        app.state.port_to_use = port_to_use

        # Read global and sensitive configuration, enable built-in TLS as needed
        ssl_kwargs = {}
        root_path = None  # Initialize root_path before try block
        try:
            from collabtrans.config.global_config import get_global_config
            from collabtrans.config.local_config import LocalConfig
            from collabtrans.config.secrets_manager import get_secrets_manager
            from collabtrans.config.global_config import save_global_config
            global_config = get_global_config()
            local_config = LocalConfig.load_from_file()

            # Get root_path for reverse proxy subpath deployment
            root_path = local_config.server.root_path if local_config.server.root_path else None
            if root_path:
                print(f"Server root_path configured: {root_path}")
                # Update FastAPI app's root_path for URL generation
                app.root_path = root_path
            
            if local_config.https.enabled:
                cert_file = local_config.https.cert_file or ""
                key_file = local_config.https.key_file or ""

                # If HTTPS is enabled but certificates are missing, try to auto-generate self-signed certificates
                if not (cert_file and key_file and os.path.exists(cert_file) and os.path.exists(key_file)):
                    try:
                        certs_dir = os.path.join(os.getcwd(), "certs")
                        os.makedirs(certs_dir, exist_ok=True)
                        default_key = os.path.join(certs_dir, "dev.local.key")
                        default_crt = os.path.join(certs_dir, "dev.local.crt")

                        if not (os.path.exists(default_key) and os.path.exists(default_crt)):
                            import subprocess
                            subprocess.run([
                                "openssl", "req", "-x509", "-nodes", "-newkey", "rsa:2048",
                                "-days", "365",
                                "-keyout", default_key,
                                "-out", default_crt,
                                "-subj", "/CN=localhost"
                            ], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                            print("Auto-generated self-signed HTTPS certificate (development environment): certs/dev.local.crt, certs/dev.local.key")

                        cert_file = default_crt
                        key_file = default_key
                        # Write back to local configuration for direct use on next startup
                        local_config.https.cert_file = cert_file
                        local_config.https.key_file = key_file
                        local_config.save_to_file()
                    except Exception as gen_err:
                        print(f"Failed to auto-generate development self-signed certificate, will start with HTTP: {gen_err}")

                if cert_file and key_file and os.path.exists(cert_file) and os.path.exists(key_file):
                    secrets_manager = get_secrets_manager()
                    key_password = None
                    try:
                        key_password = secrets_manager.get_web_tls_password()
                    except Exception:
                        key_password = None
                    ssl_kwargs.update({
                        "ssl_certfile": cert_file,
                        "ssl_keyfile": key_file,
                    })
                    if key_password:
                        ssl_kwargs["ssl_keyfile_password"] = key_password
                else:
                    print("HTTPS is enabled, but certificate or private key file does not exist, will start with HTTP.")
        except Exception as _e:
            print(f"Failed to read HTTPS configuration, will start with HTTP: {_e}")

        # Prepare uvicorn arguments
        uvicorn_kwargs = ssl_kwargs
        if root_path:
            uvicorn_kwargs["root_path"] = root_path

        uvicorn.run(app, host="0.0.0.0", port=port_to_use, workers=1, log_level="debug", access_log=False, **uvicorn_kwargs)
    except Exception as e:
        print(f"Startup failed: {e}")


if __name__ == "__main__":
    run_app()