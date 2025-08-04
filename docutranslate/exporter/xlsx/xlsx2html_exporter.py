from dataclasses import dataclass
from io import BytesIO

import jinja2
import openpyxl

from docutranslate.exporter.base import ExporterConfig
from docutranslate.exporter.xlsx.base import XlsxExporter
from docutranslate.ir.document import Document
from docutranslate.utils.resource_utils import resource_path


@dataclass
class Xlsx2HTMLExporterConfig(ExporterConfig):
    cdn: bool = True


class Xlsx2HTMLExporter(XlsxExporter):
    def __init__(self, config: Xlsx2HTMLExporterConfig = None):
        config = config or Xlsx2HTMLExporterConfig()
        super().__init__(config=config)
        self.cdn = config.cdn

    def export(self, document: Document) -> Document:

        # 1. 加载工作簿和工作表
        workbook = openpyxl.load_workbook(BytesIO(document.content))
        sheet = workbook.active

        # 2. 手动构建HTML字符串
        table = '<table border="1">\n'

        # 处理表头
        table += '  <thead>\n    <tr>\n'
        for cell in sheet[1]:  # 假设第一行是表头
            table += f'      <th>{cell.value}</th>\n'
        table += '    </tr>\n  </thead>\n'

        # 处理数据行
        table += '  <tbody>\n'
        # iter_rows(min_row=2) 从第二行开始遍历
        for row in sheet.iter_rows(min_row=2):
            table += '    <tr>\n'
            for cell in row:
                # 处理None值，防止在HTML中显示"None"
                cell_value = cell.value if cell.value is not None else ""
                table += f'      <td>{cell_value}</td>\n'
            table += '    </tr>\n'
        table += '  </tbody>\n'

        table += '</table>'

        html_template = resource_path("template/xlsx.html").read_text(encoding="utf-8")

        pico = f'<style>{resource_path("static/pico.css").read_text(encoding="utf-8")}</style>' if not self.cdn else r'<link rel="stylesheet" href="https://s4.zstatic.net/ajax/libs/picocss/2.1.1/pico.min.css" integrity="sha512-+4kjFgVD0n6H3xt19Ox84B56MoS7srFn60tgdWFuO4hemtjhySKyW4LnftYZn46k3THUEiTTsbVjrHai+0MOFw==" crossorigin="anonymous" referrerpolicy="no-referrer" />'
        render = jinja2.Template(html_template).render(
            title=document.stem,
            pico=pico,
            body=table,
        )
        return Document.from_bytes(content=render.encode("utf-8"), suffix=".html", stem=document.stem)
